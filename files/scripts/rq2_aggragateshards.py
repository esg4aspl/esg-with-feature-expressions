#!/usr/bin/env python3
"""
RQ2 Extreme Scalability Shard Aggregator
==========================================
Discovers all shard CSVs under each Case's extremeScalabilityTestPipeline/
directory, merges them per approach, and produces:

  - RQ2_rawData_<CaseName>.xlsx     (all shards concatenated, one sheet per approach)
  - RQ2_aggregated_<CaseName>.xlsx  (aggregated per RunID, one sheet per approach)

Usage:
    python3 rq2_aggregateshards.py
"""

import pandas as pd
import numpy as np
from pathlib import Path
import sys
from typing import Dict, List, Tuple, Optional
from collections import OrderedDict
import warnings
warnings.filterwarnings('ignore')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

SPL_NAME_MAPPING = {
    "SodaVendingMachine": "SVM",
    "eMail": "eM",
    "Elevator": "El",
    "BankAccountv2": "BAv2",
    "StudentAttendanceSystem": "SAS",
    "Tesla": "Te",
    "syngovia": "Svia",
    "HockertyShirts": "HS"
}

# Canonical sheet ordering
APPROACH_ORDER = [
    "ESG-Fx_L1", "ESG-Fx_L2", "ESG-Fx_L3", "ESG-Fx_L4",
    "EFG_L2", "EFG_L3", "EFG_L4",
    "RandomWalk_L0"
]

# Header styling constants
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='Arial', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


# =============================================================================
# AGGREGATION RULES PER APPROACH
# =============================================================================

ESGFX_L1_RULES = {
    'run_id_col': 'RunID',
    'key': ['Run ID', 'SPL Name', 'Coverage Type'],
    'sum': [
        'Total Elapsed Time(ms)', 'SAT Time(ms)', 'Product Gen Time(ms)',
        'Test Generation Time(ms)',
        'Number of ESGFx Vertices', 'Number of ESGFx Edges',
        'Number of ESGFx Test Cases', 'Number of ESGFx Test Events',
        'Coverage Analysis Time(ms)',
        'Test Execution Time(ms)',
        'Processed Products', 'Failed Products'
    ],
    'max': ['Test Generation Peak Memory(MB)', 'Test Execution Peak Memory(MB)'],
    'weighted_avg': {'L1 Coverage(%)': 'Processed Products'},
}

ESGFX_L234_RULES = {
    'run_id_col': 'RunID',
    'key': ['RunID', 'SPL Name', 'Coverage Type'],
    'sum': [
        'Total Elapsed Time(ms)', 'SAT Time(ms)', 'Product Gen Time(ms)',
        'Transformation Time(ms)', 'Test Generation Time(ms)',
        'Number of ESGFx Vertices', 'Number of ESGFx Edges',
        'Number of ESGFx Test Cases', 'Number of ESGFx Test Events',
        'Coverage Analysis Time(ms)',
        'Test Execution Time(ms)',
        'Processed Products', 'Failed Products'
    ],
    'max': ['Test Generation Peak Memory(MB)', 'Test Execution Peak Memory(MB)'],
    # Coverage column name varies: L2 Coverage(%), L3 Coverage(%), L4 Coverage(%)
    'weighted_avg': {},  # filled dynamically
}

EFG_RULES = {
    'run_id_col': 'RunID',
    'key': ['RunID', 'SPL Name', 'Coverage Type'],
    'sum': [
        'Total Elapsed Time(ms)', 'SAT Time(ms)', 'Product Gen Time(ms)',
        'EFG Transformation Time(ms)', 'Test Generation Time(ms)', 'Parse Time(ms)',
        'Number of EFG Vertices', 'Number of EFG Edges',
        'Number of EFG Test Cases', 'Number of EFG Test Events',
        'Event Coverage Analysis Time(ms)', 'Edge Coverage Analysis Time(ms)',
        'Test Execution Time(ms)',
        'Processed Products', 'Failed Products'
    ],
    'max': ['Test Generation Peak Memory(MB)', 'Test Execution Peak Memory(MB)'],
    'weighted_avg': {
        'Event Coverage(%)': 'Processed Products',
        'Edge Coverage(%)': 'Processed Products',
    },
}

RW_RULES = {
    'run_id_col': 'RunID',
    'key': ['RunID', 'SPL Name', 'Coverage Type'],
    'sum': [
        'Total Elapsed Time(ms)', 'SAT Time(ms)', 'Product Gen Time(ms)',
        'Test Generation Time(ms)',
        'Number of Vertices', 'Number of Edges',
        'Number of Test Cases', 'Number of Test Events', 'Aborted Sequences',
        'Event Coverage Analysis Time(ms)', 'Edge Coverage Analysis Time(ms)',
        'Test Execution Time(ms)',
        'Safety Limit Hit Count',
        'Processed Products', 'Failed Products'
    ],
    'max': ['Test Generation Peak Memory(MB)', 'Test Execution Peak Memory(MB)'],
    'weighted_avg': {
        'Event Coverage(%)': 'Processed Products',
        'Edge Coverage(%)': 'Processed Products',
        'Avg Time on Safety Limit(ms)': 'Safety Limit Hit Count',
        'Avg Steps on Safety Limit': 'Safety Limit Hit Count',
        'Avg Coverage at Safety Limit(%)': 'Safety Limit Hit Count',
    },
}


def get_rules_for_approach(approach_key: str, df_columns: list) -> dict:
    """Return the appropriate aggregation rules for a given approach key."""
    if approach_key == 'ESG-Fx_L1':
        return ESGFX_L1_RULES
    elif approach_key.startswith('ESG-Fx_L'):
        level = approach_key.split('_L')[1]
        rules = {k: (v.copy() if isinstance(v, (list, dict)) else v)
                 for k, v in ESGFX_L234_RULES.items()}
        cov_col = f'L{level} Coverage(%)'
        rules['weighted_avg'] = {cov_col: 'Processed Products'}
        return rules
    elif approach_key.startswith('EFG_L'):
        return EFG_RULES
    elif approach_key == 'RandomWalk_L0':
        return RW_RULES
    return {}


# =============================================================================
# CORE AGGREGATION FUNCTION
# =============================================================================

def aggregate_by_run_id(df: pd.DataFrame, rules: dict) -> pd.DataFrame:
    """Aggregate shard rows by Run ID according to approach-specific rules."""
    if df.empty:
        return df

    run_id_col = rules.get('run_id_col', 'RunID')
    if run_id_col not in df.columns:
        # Try alternative
        alt = 'Run ID' if run_id_col == 'RunID' else 'RunID'
        if alt in df.columns:
            run_id_col = alt
        else:
            print(f"    WARNING: No run ID column found in columns: {list(df.columns)}")
            return df

    grouped = df.groupby(run_id_col, sort=False)
    aggregated_rows = []

    for run_id, group in grouped:
        row = {}

        # Key columns: take first value
        for col in rules.get('key', []):
            if col in group.columns:
                row[col] = group[col].iloc[0]

        # Sum columns
        for col in rules.get('sum', []):
            if col in group.columns:
                row[col] = group[col].sum()

        # Max columns
        for col in rules.get('max', []):
            if col in group.columns:
                row[col] = group[col].max()

        # Weighted average columns
        for col, weight_col in rules.get('weighted_avg', {}).items():
            if col in group.columns and weight_col in group.columns:
                weights = group[weight_col]
                values = group[col]
                total_weight = weights.sum()
                if total_weight > 0:
                    row[col] = (values * weights).sum() / total_weight
                else:
                    row[col] = 0.0

        aggregated_rows.append(row)

    result = pd.DataFrame(aggregated_rows)

    # Preserve original column order
    ordered_cols = [c for c in df.columns if c in result.columns]
    extra_cols = [c for c in result.columns if c not in ordered_cols]
    return result[ordered_cols + extra_cols]


# =============================================================================
# PROJECT ROOT FINDER
# =============================================================================

def find_project_root() -> Optional[Path]:
    """Find project root by looking for 'files/Cases' directory."""
    candidates = [
        Path.cwd(),
        Path.cwd().parent,
        Path.cwd().parent.parent,
    ]
    # Also check if we're inside files/scripts
    if Path.cwd().name == "scripts" and Path.cwd().parent.name == "files":
        candidates.append(Path.cwd().parent.parent)

    for candidate in candidates:
        if (candidate / "files" / "Cases").exists():
            return candidate
    return None


# =============================================================================
# SHARD FILE DISCOVERY
# =============================================================================

def discover_shards(pipeline_dir: Path, spl_folder_name: str) -> Dict[str, List[Path]]:
    """
    Discover all shard CSV files under the extremeScalabilityTestPipeline directory.

    Expected structure:
        pipeline_dir/
            ESG-Fx/L1/<prefix>_ESG-Fx_L1_shard*.csv
            ESG-Fx/L2/<prefix>_ESG-Fx_L2_shard*.csv
            ...
            EFG/L2/<prefix>_EFG_L2_shard*.csv
            ...
            RandomWalk/L0/<prefix>_RandomWalk_L0_shard*.csv

    Returns dict: approach_key -> list of shard file Paths
    """
    file_prefix = SPL_NAME_MAPPING.get(spl_folder_name, spl_folder_name)
    results = OrderedDict()

    search_specs = [
        ("ESG-Fx_L1", "ESG-Fx/L1", "_ESG-Fx_L1_shard"),
        ("ESG-Fx_L2", "ESG-Fx/L2", "_ESG-Fx_L2_shard"),
        ("ESG-Fx_L3", "ESG-Fx/L3", "_ESG-Fx_L3_shard"),
        ("ESG-Fx_L4", "ESG-Fx/L4", "_ESG-Fx_L4_shard"),
        ("EFG_L2",    "EFG/L2",    "_EFG_L2_shard"),
        ("EFG_L3",    "EFG/L3",    "_EFG_L3_shard"),
        ("EFG_L4",    "EFG/L4",    "_EFG_L4_shard"),
        ("RandomWalk_L0", "RandomWalk/L0", "_RandomWalk_L0_shard"),
    ]

    for approach_key, subdir, pattern_suffix in search_specs:
        full_dir = pipeline_dir / subdir
        if not full_dir.exists():
            continue

        # Try file prefix first, then folder name
        for prefix in [file_prefix, spl_folder_name]:
            pattern = f"{prefix}{pattern_suffix}*.csv"
            files = sorted(full_dir.glob(pattern))
            if files:
                results[approach_key] = files
                break

    return results


# =============================================================================
# EXCEL FORMATTING
# =============================================================================

def format_workbook(wb_path: Path):
    """Apply professional formatting to all sheets in a workbook."""
    wb = load_workbook(wb_path)

    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue

        # Style header row
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Style data rows
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-fit column widths
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

        # Freeze header row
        ws.freeze_panes = 'A2'

    wb.save(wb_path)


# =============================================================================
# PROCESS SINGLE SPL
# =============================================================================

def process_spl(spl_folder_name: str, pipeline_dir: Path, output_dir: Path) -> bool:
    """Process all RQ2 shards for a single SPL case."""
    file_prefix = SPL_NAME_MAPPING.get(spl_folder_name, spl_folder_name)

    print(f"\n{'='*80}")
    print(f"  PROCESSING: {spl_folder_name} (file prefix: {file_prefix})")
    print(f"{'='*80}")
    print(f"  Pipeline dir: {pipeline_dir}")

    shards = discover_shards(pipeline_dir, spl_folder_name)

    if not shards:
        print("  WARNING: No shard files found")
        return False

    raw_sheets = OrderedDict()
    agg_sheets = OrderedDict()

    for approach_key in APPROACH_ORDER:
        if approach_key not in shards:
            continue

        shard_files = shards[approach_key]
        print(f"\n  [{approach_key}] Found {len(shard_files)} shards")

        # Read and concatenate all shards
        dfs = []
        for f in shard_files:
            try:
                df = pd.read_csv(f, sep=';', decimal=',')
                dfs.append(df)
            except Exception as e:
                print(f"    ERROR reading {f.name}: {e}")
                continue

        if not dfs:
            print(f"    WARNING: No valid data for {approach_key}")
            continue

        raw_df = pd.concat(dfs, ignore_index=True)
        print(f"    Raw: {len(raw_df)} rows")

        # Aggregate by RunID
        rules = get_rules_for_approach(approach_key, list(raw_df.columns))
        agg_df = aggregate_by_run_id(raw_df, rules)
        print(f"    Aggregated: {len(agg_df)} runs")

        # Compute throughput per run for the aggregated data
        run_id_col = rules.get('run_id_col', 'RunID')
        if run_id_col not in agg_df.columns:
            run_id_col = 'Run ID' if 'Run ID' in agg_df.columns else 'RunID'

        if 'Total Elapsed Time(ms)' in agg_df.columns and 'Processed Products' in agg_df.columns:
            time_hours = agg_df['Total Elapsed Time(ms)'] / 3_600_000
            # Avoid division by zero
            agg_df['Throughput (products/hour)'] = np.where(
                time_hours > 0,
                agg_df['Processed Products'] / time_hours,
                0
            )
            median_tp = agg_df['Throughput (products/hour)'].median()
            total_products = agg_df['Processed Products'].iloc[0] if len(agg_df) > 0 else 0
            median_time = agg_df['Total Elapsed Time(ms)'].median()
            print(f"    Products/run: {total_products:.0f}")
            print(f"    Median time: {median_time:,.0f} ms ({median_time/3_600_000:.2f} h)")
            print(f"    Median throughput: {median_tp:,.0f} products/hour")

        raw_sheets[approach_key] = raw_df
        agg_sheets[approach_key] = agg_df

    if not raw_sheets:
        print("  WARNING: No results to save")
        return False

    # Write Excel files
    output_dir.mkdir(parents=True, exist_ok=True)

    raw_path = output_dir / f"RQ2_rawData_{spl_folder_name}.xlsx"
    agg_path = output_dir / f"RQ2_aggregated_{spl_folder_name}.xlsx"

    print(f"\n  Saving raw data: {raw_path.name}")
    with pd.ExcelWriter(raw_path, engine='openpyxl') as writer:
        for sheet_name, df in raw_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"  Saving aggregated data: {agg_path.name}")
    with pd.ExcelWriter(agg_path, engine='openpyxl') as writer:
        for sheet_name, df in agg_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Apply formatting
    format_workbook(raw_path)
    format_workbook(agg_path)

    print(f"  OK: {len(raw_sheets)} approach sheets written")
    return True


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 80)
    print("RQ2 EXTREME SCALABILITY - SHARD AGGREGATION")
    print("=" * 80)

    project_root = find_project_root()

    if not project_root:
        print("ERROR: Could not find project root (looking for files/Cases)")
        sys.exit(1)

    cases_dir = project_root / "files" / "Cases"
    print(f"Project root: {project_root}")
    print(f"Cases dir:    {cases_dir}")

    # Discover SPL directories
    spls = sorted([d for d in cases_dir.iterdir() if d.is_dir()])

    if not spls:
        print("ERROR: No SPL directories found!")
        sys.exit(1)

    print(f"\nFound {len(spls)} SPL directories:")
    for spl_dir in spls:
        prefix = SPL_NAME_MAPPING.get(spl_dir.name, spl_dir.name)
        marker = f" -> {prefix}" if prefix != spl_dir.name else ""
        print(f"  - {spl_dir.name}{marker}")

    processed = 0
    skipped = 0
    failed = 0

    for spl_dir in spls:
        pipeline_dir = spl_dir / "extremeScalabilityTestPipeline"

        if not pipeline_dir.exists():
            print(f"\n  Skipping {spl_dir.name}: no extremeScalabilityTestPipeline/")
            skipped += 1
            continue

        try:
            success = process_spl(spl_dir.name, pipeline_dir, pipeline_dir)
            if success:
                processed += 1
            else:
                failed += 1
        except Exception as e:
            print(f"\n  ERROR processing {spl_dir.name}: {e}")
            import traceback
            traceback.print_exc()
            failed += 1

    print(f"\n{'='*80}")
    print("SUMMARY")
    print(f"{'='*80}")
    print(f"  Processed: {processed}")
    print(f"  Skipped:   {skipped}")
    if failed:
        print(f"  Failed:    {failed}")
    print(f"\nOutput per case in extremeScalabilityTestPipeline/:")
    print(f"  - RQ2_rawData_<CaseName>.xlsx")
    print(f"  - RQ2_aggregated_<CaseName>.xlsx")
    print("=" * 80)


if __name__ == "__main__":
    main()