#!/usr/bin/env python3
"""
rq5_main.py — Orchestrator for RQ5 Test Suite Redundancy Analysis
===================================================================
Reads data using rq5_readers, computes metrics using rq5_analysis,
generates plots using rq5_plots, and writes Excel reports.

Usage:
    python3 rq5_main.py <base_path> <spl_name> <config_filename> [approaches...]

Examples:
    # All ESG-Fx approaches (default):
    python3 rq5_main.py files/Cases/Elevator El El_ProductConfigurations.txt

    # Specific approaches:
    python3 rq5_main.py files/Cases/syngovia Svia Svia_ProductConfigurations.txt ESG-Fx_L1 ESG-Fx_L2 EFG_L2

    # All available (ESG-Fx + EFG):
    python3 rq5_main.py files/Cases/syngovia Svia Svia_ProductConfigurations.txt ALL

Output:
    <base_path>/testsuiteanalysis/
        ├── <SPL>_RQ5_Pairwise_<approach>.xlsx    (per approach: pairwise data + correlations)
        ├── <SPL>_RQ5_Summary.xlsx                (cross-approach summary)
        ├── <SPL>_RQ5_Scatter_<approach>.pdf      (per approach: JSS scatter plots)
        ├── <SPL>_RQ5_ScatterOverview.png         (combined overview)
        ├── <SPL>_RQ5_CumulativeEdge.pdf          (cumulative edge coverage)
        └── <SPL>_RQ5_UniqueRatio.pdf             (unique sequence ratio box plot)
"""

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from rq5_readers import read_product_configurations, discover_all_suites
from rq5_analysis import (
    analyze_approach,
    compute_unique_sequence_ratios,
    compute_cumulative_edge_coverage,
)
from rq5_plots import (
    plot_scatter_per_approach,
    plot_cumulative_edge_coverage,
    plot_unique_sequence_ratios,
)


# =============================================================================
# EXCEL STYLING
# =============================================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
CELL_FONT = Font(name='Arial', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _style_header(ws, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def _style_data(ws, n_rows, n_cols):
    for row in range(2, n_rows + 2):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if isinstance(cell.value, float):
                cell.number_format = '0.0000'


def _auto_width(ws, df):
    for idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), 10)
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = max_len + 3
    ws.freeze_panes = 'A2'


# =============================================================================
# EXCEL WRITERS
# =============================================================================

def write_pairwise_excel(df, correlations, output_path):
    """Write per-approach pairwise results to Excel."""
    wb = Workbook()

    # Sheet 1: Pairwise data
    ws1 = wb.active
    ws1.title = 'Pairwise'
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    _style_header(ws1, len(df.columns))
    _style_data(ws1, len(df), len(df.columns))
    _auto_width(ws1, df)

    # Sheet 2: Correlations
    ws2 = wb.create_sheet('Correlations')
    ws2.append(['Similarity_Type', 'Spearman_rho', 'Spearman_p', 'Pearson_r', 'Pearson_p', 'N_pairs'])
    for metric, vals in correlations.items():
        ws2.append([metric, vals['spearman_rho'], vals['spearman_p'],
                     vals['pearson_r'], vals['pearson_p'], vals['n']])
    _style_header(ws2, 6)

    # Sheet 3: Descriptive statistics
    ws3 = wb.create_sheet('Descriptive')
    desc = df[['Config_Jaccard', 'Sequence_Jaccard', 'Edge_Jaccard']].describe().reset_index()
    for r in dataframe_to_rows(desc, index=False, header=True):
        ws3.append(r)
    _style_header(ws3, len(desc.columns))

    wb.save(output_path)


def write_summary_excel(all_results, unique_ratios_all, cumulative_all, output_path, spl_name):
    """Write cross-approach summary to Excel."""
    wb = Workbook()

    # Sheet 1: Correlation comparison
    ws1 = wb.active
    ws1.title = 'Correlations'
    headers = ['Approach', 'Similarity_Type', 'Spearman_rho', 'Spearman_p', 'Pearson_r', 'Pearson_p', 'N_pairs']
    ws1.append(headers)
    for approach, data in all_results.items():
        for metric, vals in data['correlations'].items():
            ws1.append([approach, metric, vals['spearman_rho'], vals['spearman_p'],
                         vals['pearson_r'], vals['pearson_p'], vals['n']])
    _style_header(ws1, len(headers))

    # Sheet 2: Descriptive stats per approach
    ws2 = wb.create_sheet('DescriptiveStats')
    ws2.append(['Approach', 'Metric', 'Mean', 'Median', 'Std', 'Min', 'Max', 'N_products', 'N_pairs'])
    for approach, data in all_results.items():
        df = data['dataframe']
        for col in ['Config_Jaccard', 'Sequence_Jaccard', 'Edge_Jaccard']:
            ws2.append([approach, col, df[col].mean(), df[col].median(), df[col].std(),
                         df[col].min(), df[col].max(), data['n_products'], data['n_pairs']])
    _style_header(ws2, 9)

    # Sheet 3: Unique Sequence Ratios
    ws3 = wb.create_sheet('UniqueSequenceRatio')
    ws3.append(['Approach', 'Product', 'Unique_Sequences', 'Total_Sequences', 'Unique_Ratio'])
    for approach, ratios in unique_ratios_all.items():
        for product, (unique, total, ratio) in sorted(ratios.items()):
            ws3.append([approach, product, unique, total, ratio])
    _style_header(ws3, 5)

    # Sheet 4: Cumulative Edge Coverage
    ws4 = wb.create_sheet('CumulativeEdgeCoverage')
    ws4.append(['Approach', 'Product_Index', 'Cumulative_Unique_Edges'])
    for approach, curve in cumulative_all.items():
        for idx, edges in curve:
            ws4.append([approach, idx, edges])
    _style_header(ws4, 3)

    wb.save(output_path)


# =============================================================================
# MAIN
# =============================================================================

DEFAULT_APPROACHES = ['ESG-Fx_L0', 'ESG-Fx_L1', 'ESG-Fx_L2', 'ESG-Fx_L3', 'ESG-Fx_L4']
ALL_APPROACHES = DEFAULT_APPROACHES + ['EFG_L2', 'EFG_L3', 'EFG_L4']


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)

    base_path = sys.argv[1]
    spl_name = sys.argv[2]
    config_filename = sys.argv[3]
    
    # Parse approach list
    if len(sys.argv) > 4:
        if sys.argv[4] == 'ALL':
            approaches = ALL_APPROACHES
        else:
            approaches = sys.argv[4:]
    else:
        approaches = DEFAULT_APPROACHES

    config_path = os.path.join(base_path, config_filename)

    print(f"{'=' * 60}")
    print(f"RQ5 REDUNDANCY ANALYSIS: {spl_name}")
    print(f"Base path:  {base_path}")
    print(f"Config:     {config_path}")
    print(f"Approaches: {', '.join(approaches)}")
    print(f"{'=' * 60}")

    # 1. Read configurations
    configs = read_product_configurations(config_path)
    print(f"\nLoaded {len(configs)} product configurations.")

    # 2. Create output directory
    output_dir = os.path.join(base_path, 'testsuiteanalysis')
    os.makedirs(output_dir, exist_ok=True)

    # 3. Analyze each approach
    all_results = {}
    unique_ratios_all = {}
    cumulative_all = {}

    for approach in approaches:
        print(f"\n{'─' * 40}")
        print(f"Loading: {approach}")
        suites = discover_all_suites(base_path, approach)
        
        if len(suites) < 2:
            print(f"  Found {len(suites)} suites — skipping (need ≥ 2).")
            continue

        print(f"  Found {len(suites)} product test suites.")
        result = analyze_approach(configs, suites, approach)
        
        if result:
            all_results[approach] = result
            unique_ratios_all[approach] = result['unique_ratios']
            cumulative_all[approach] = result['cumulative_curve']

            # Write per-approach Excel
            safe = approach.replace('-', '').replace('_', '')
            excel_path = os.path.join(output_dir, f'{spl_name}_RQ5_Pairwise_{safe}.xlsx')
            write_pairwise_excel(result['dataframe'], result['correlations'], excel_path)
            print(f"  Excel: {excel_path}")

    if not all_results:
        print("\nNo approaches had sufficient data. Exiting.")
        sys.exit(1)

    # 4. Cross-approach summary Excel
    print(f"\n{'─' * 40}")
    print("Writing cross-approach summary...")
    summary_path = os.path.join(output_dir, f'{spl_name}_RQ5_Summary.xlsx')
    write_summary_excel(all_results, unique_ratios_all, cumulative_all, summary_path, spl_name)
    print(f"  Summary: {summary_path}")

    # 5. Generate plots
    print(f"\nGenerating plots...")
    plot_scatter_per_approach(all_results, output_dir, spl_name)
    plot_cumulative_edge_coverage(cumulative_all, output_dir, spl_name)
    plot_unique_sequence_ratios(unique_ratios_all, output_dir, spl_name)

    # 6. Final summary
    print(f"\n{'=' * 60}")
    print(f"COMPLETE: {spl_name}")
    print(f"{'=' * 60}")
    for approach, data in all_results.items():
        df = data['dataframe']
        corr = data['correlations']
        print(f"\n  {approach} ({data['n_products']} products, {data['n_pairs']} pairs):")
        print(f"    Config Jaccard:    mean={df['Config_Jaccard'].mean():.3f}  "
              f"[{df['Config_Jaccard'].min():.3f} — {df['Config_Jaccard'].max():.3f}]")
        for metric in ['Sequence_Jaccard', 'Edge_Jaccard']:
            c = corr[metric]
            print(f"    {metric:20s}: ρ={c['spearman_rho']:+.4f}  r={c['pearson_r']:+.4f}  "
                  f"mean={df[metric].mean():.3f}")
        ur = [r[2] for r in data['unique_ratios'].values()]
        print(f"    Unique Seq Ratio:  mean={sum(ur)/len(ur):.3f}")

    print(f"\n  Output: {output_dir}/")


if __name__ == '__main__':
    main()