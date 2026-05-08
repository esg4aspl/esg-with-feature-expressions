#!/usr/bin/env python3
"""
rq2_06_time_breakdown.py  (RQ2 Step 6 — KEY ANALYSIS)

Decomposes the end-to-end RQ2 pipeline cost for each (SPL x approach)
into its component phases and reports two share metrics that drive the
manuscript narrative:

  SATShare(%)      -- share of T_pipeline spent in SAT solving.
  OwnCostShare(%)  -- share of T_pipeline that is NOT SAT or
                      ProductGen, i.e. the per-product overhead each
                      approach adds on top of the shared infrastructure.

Why these two matter
--------------------
At small/medium SPLs, ProductGen + the approach-specific phases
(Transformation + TestGen + TestExec) dominate the pipeline; throughput
gaps between approaches are large. At industrial SPLs (Tesla,
HockertyShirts), SAT solving consumes 99%+ of T_pipeline -- all three
approaches pay the same SAT bill, so end-to-end throughput converges.
"OwnCost" isolates the part of the pipeline that the approach actually
controls and supports the manuscript's claim:
"Model Once, Generate Any reduces per-product overhead on top of the
shared SAT cost".

Inputs
------
Reads the canonical per-shard schema produced by Step 1:
    files/Cases/<SPL>/RQ2_perShard_<SPL>.xlsx
Aggregation: per-shard median across runs, then sum across the 80
shards. This is the same SERIAL aggregation used in
RQ2_SPLSummary_medians.xlsx so the numbers reconcile.

Pipeline definition: T_pipeline = SatTime + ProdGenTime + TestGenTime
                                  + TestExecTime
                                  ( + EFGTransformationTime + ParsingTime for EFG )
EXCLUDES: TransformationTime for ESG-Fx (already inside TestGenTime --
          a Java instrumentation bug); coverage-analysis times
          (validation observers, not deployment cost). The audit-only
          'Java_T_total_buggy(ms)' is NOT used.

Output
------
files/scripts/statistical_test_scripts/rq2_result/rq2_time_breakdown.xlsx

  - 'breakdown'           : full per-(SPL x approach x level) table
  - 'sat_share_pivot'     : SPL rows x Approach cols, SatShare(%)
  - 'own_cost_share_pivot': SPL rows x Approach cols, OwnCostShare(%)

Plus stacked-bar PDF in plots/rq2_time_breakdown.pdf.
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# ─── Configuration ─────────────────────────────────────────────────────
SPL_MAPPING = {
    "SodaVendingMachine":      "SVM",
    "eMail":                   "eM",
    "Elevator":                "El",
    "BankAccountv2":           "BAv2",
    "StudentAttendanceSystem": "SAS",
    "syngovia":                "Svia",
    "Tesla":                   "Te",
    "HockertyShirts":          "HS",
}
SPL_ORDER = ["SVM", "eM", "El", "BAv2", "SAS", "Svia", "Te", "HS"]

# Approach configurations for each coverage level. The DEFAULT (L=2) is
# the article-scope set used to generate the main manuscript tables.
# To produce the L=3 or L=4 supplementary breakdowns, run with:
#     python rq2_06_time_breakdown.py --levels L3
#     python rq2_06_time_breakdown.py --levels L4
# The output Excel and PDF filenames carry the level suffix, so existing
# L=2 outputs are not overwritten.
APPROACH_CONFIG_BY_LEVEL = {
    "L2": {
        "ESG-Fx_L2":     "Model Once, Generate Any (L=2)",
        "EFG_L2":        "Structural Baseline (L=2)",
        "RandomWalk_L0": "Stochastic Baseline",
    },
    "L3": {
        "ESG-Fx_L3":     "Model Once, Generate Any (L=3)",
        "EFG_L3":        "Structural Baseline (L=3)",
        "RandomWalk_L0": "Stochastic Baseline",
    },
    "L4": {
        "ESG-Fx_L4":     "Model Once, Generate Any (L=4)",
        "EFG_L4":        "Structural Baseline (L=4)",
        "RandomWalk_L0": "Stochastic Baseline",
    },
}

# Backwards-compatible default: L=2 is what the manuscript main tables use.
APPROACH_CONFIG = APPROACH_CONFIG_BY_LEVEL["L2"]

COARSE_PHASES = ["SAT", "ProductGen", "Transformation", "TestGen", "TestExec", "Parsing", "Other"]
PHASE_COLORS = {
    "SAT":            "#d95f02",
    "ProductGen":     "#1b9e77",
    "Transformation": "#7570b3",
    "TestGen":        "#e7298a",
    "TestExec":       "#66a61e",
    "Parsing":        "#a6761d",
    "Other":          "#999999",
}


# ─── Path discovery (no absolute paths) ────────────────────────────────
def find_project_root():
    candidates = [Path.cwd()]
    candidates.extend(Path.cwd().parents)
    candidates.append(Path(__file__).resolve().parent)
    candidates.extend(Path(__file__).resolve().parents)
    seen = set()
    for c in candidates:
        c = c.resolve()
        if c in seen:
            continue
        seen.add(c)
        if (c / "files" / "Cases").exists():
            return c
    return None


# ─── Per-cell aggregation ──────────────────────────────────────────────
def aggregate_one_cell(perShard_path, sheet_name):
    """Read one (SPL, approach x level) sheet from a per-shard Excel,
    take per-shard median across runs, then sum across shards.

    Returns a dict of phase totals (in ms) and total HandledProducts,
    or None if the sheet doesn't exist or has no data.
    """
    try:
        df = pd.read_excel(perShard_path, sheet_name=sheet_name)
    except Exception:
        return None
    if df.empty or "Shard" not in df.columns:
        return None

    # Per-shard median across runs (the standard RQ2 aggregation step;
    # done in rq2_02 normally but recomputed here so this script can
    # operate directly on perShard data without depending on Step 2's
    # output existing).
    numeric = df.select_dtypes(include=[np.number])
    per_shard = df.groupby("Shard")[list(numeric.columns)].median()

    # Skip shards where no products were processed (incomplete runs)
    if "HandledProducts" in per_shard.columns:
        per_shard = per_shard[per_shard["HandledProducts"] > 0]
    if per_shard.empty:
        return None

    out = {
        "T_pipeline(ms)":             float(per_shard.get("T_pipeline(ms)", pd.Series(dtype=float)).sum()),
        "SatTime(ms)":                float(per_shard.get("SatTime(ms)", pd.Series(dtype=float)).sum()),
        "ProdGenTime(ms)":            float(per_shard.get("ProdGenTime(ms)", pd.Series(dtype=float)).sum()),
        "TransformationTime(ms)":     float(per_shard.get("TransformationTime(ms)", pd.Series(dtype=float)).sum()),
        "EFGTransformationTime(ms)":  float(per_shard.get("EFGTransformationTime(ms)", pd.Series(dtype=float)).sum()),
        "TestGenTime(ms)":            float(per_shard.get("TestGenTime(ms)", pd.Series(dtype=float)).sum()),
        "ParsingTime(ms)":            float(per_shard.get("ParsingTime(ms)", pd.Series(dtype=float)).sum()),
        "TestExecTime(ms)":           float(per_shard.get("TestExecTime(ms)", pd.Series(dtype=float)).sum()),
        "HandledProducts":            int(per_shard.get("HandledProducts", pd.Series(dtype=float)).sum()),
    }
    return out


def coarse_phases(totals, approach_label):
    """Map fine-grained columns to the 7 coarse phases for plotting.

    Note that for ESG-Fx the TransformationTime is intentionally
    excluded from T_pipeline (it is double-counted inside TestGenTime
    in the Java instrumentation, so adding it again would inflate the
    pipeline). It IS reported in the breakdown table as 'Transformation'
    for transparency, but T_pipeline = SAT + ProductGen + TestGen +
    TestExec only -- so the 'Transformation' value here is for
    informational display, NOT a separate share of pipeline time.
    """
    sat = totals["SatTime(ms)"]
    prod = totals["ProdGenTime(ms)"]
    tg = totals["TestGenTime(ms)"]
    te = totals["TestExecTime(ms)"]
    if approach_label.startswith("EFG"):
        # EFG has its own transformation step that IS part of T_pipeline
        # (the conversion from product .EFG to internal representation).
        trans = totals["EFGTransformationTime(ms)"]
        parsing = totals["ParsingTime(ms)"]
    else:
        trans = 0.0
        parsing = 0.0
    return {
        "SAT":            sat,
        "ProductGen":     prod,
        "Transformation": trans,
        "TestGen":        tg,
        "TestExec":       te,
        "Parsing":        parsing,
    }


def build_breakdown_row(spl_short, sheet_name, approach_label, totals):
    """Compute breakdown + share columns for one cell."""
    t_pipe = totals["T_pipeline(ms)"]
    if t_pipe <= 0:
        return None

    phases = coarse_phases(totals, sheet_name)
    sat = phases["SAT"]
    prod = phases["ProductGen"]

    # Sum of named phases excluding 'Other'
    named = phases["SAT"] + phases["ProductGen"] + phases["Transformation"] + \
            phases["TestGen"] + phases["TestExec"] + phases["Parsing"]
    other = max(0.0, t_pipe - named)

    own_cost = max(0.0, t_pipe - sat - prod)

    return {
        "SPL":                       spl_short,
        "Sheet":                     sheet_name,
        "Approach":                  approach_label,
        "HandledProducts":           totals["HandledProducts"],
        "T_pipeline(ms)":            round(t_pipe, 2),
        "T_pipeline(hours)":         round(t_pipe / 1000 / 3600, 3),
        "SatTime(ms)":               round(sat, 2),
        "SATShare(%)":               round(100 * sat / t_pipe, 2),
        "ProdGenTime(ms)":           round(prod, 2),
        "ProdGenShare(%)":           round(100 * prod / t_pipe, 2),
        "TransformationTime(ms)":    round(phases["Transformation"], 2),
        "TestGenTime(ms)":           round(phases["TestGen"], 2),
        "TestGenShare(%)":           round(100 * phases["TestGen"] / t_pipe, 2),
        "ParsingTime(ms)":           round(phases["Parsing"], 2),
        "TestExecTime(ms)":          round(phases["TestExec"], 2),
        "Other(ms)":                 round(other, 2),
        "OwnCost(ms)":               round(own_cost, 2),
        "OwnCostShare(%)":           round(100 * own_cost / t_pipe, 2),
        # Keep the raw phase values for the stacked-bar plot
        "_phases":                   phases,
        "_other":                    other,
    }


# ─── openpyxl post-processing: auto-fit + nice headers ─────────────────
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FONT = Font(bold=True, name='Calibri', size=11)
HEADER_FILL = PatternFill('solid', fgColor='E0E0E0')


def auto_fit_workbook(path, max_width=42):
    """Open an .xlsx written by pandas, autosize columns to header
    width (with a sensible cap), and apply a header style. The cap
    avoids comically wide columns when one cell happens to hold a
    long string."""
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        # First row = header (pandas writes one). Style it.
        for col_idx in range(1, ws.max_column + 1):
            c = ws.cell(1, col_idx)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.border = THIN
            c.alignment = Alignment(horizontal='left', vertical='center')
        # Width: take the MAX of header length and longest data value
        # in the column (capped). Header alone often beats data, since
        # 'TestGenPeakMemory(MB)' > '128.0'.
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(1, col_idx).value
            best = len(str(header_val)) if header_val else 8
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, col_idx).value
                if v is not None:
                    best = max(best, len(str(v)))
            width = min(best * 1.1 + 2, max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)


# ─── Main ──────────────────────────────────────────────────────────────
def main():
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "Decompose RQ2 pipeline cost by phase and report SAT/OwnCost "
            "shares. Default operates on L=2 (manuscript main scope); "
            "use --levels L3 or --levels L4 for the supplementary "
            "breakdowns."
        )
    )
    parser.add_argument(
        "--levels",
        nargs="+",
        default=["L2"],
        choices=list(APPROACH_CONFIG_BY_LEVEL.keys()),
        help="Coverage level(s) to process. Default: L2.",
    )
    args = parser.parse_args()

    print("=" * 72)
    print("rq2_06 -- Time Breakdown / SATShare / OwnCostShare")
    print(f"Coverage levels: {', '.join(args.levels)}")
    print("=" * 72)

    project_root = find_project_root()
    if not project_root:
        print("ERROR: could not find project root containing files/Cases/.")
        sys.exit(1)

    cases_dir = project_root / "files" / "Cases"
    out_dir = project_root / "files" / "scripts" / "statistical_test_scripts" / "rq2_result"
    plots_dir = out_dir / "plots"
    out_dir.mkdir(parents=True, exist_ok=True)
    plots_dir.mkdir(parents=True, exist_ok=True)
    print(f"Project root: {project_root}")
    print(f"Output dir  : {out_dir}\n")

    # Process each requested level into its own Excel + PDF.
    for level in args.levels:
        approach_config = APPROACH_CONFIG_BY_LEVEL[level]
        process_level(
            level=level,
            approach_config=approach_config,
            cases_dir=cases_dir,
            out_dir=out_dir,
            plots_dir=plots_dir,
            project_root=project_root,
        )

    print("\nrq2_06 DONE.")


def process_level(level, approach_config, cases_dir, out_dir, plots_dir,
                  project_root):
    """Produce the breakdown Excel and stacked-bar PDF for one level."""
    print(f"\n--- Processing {level} ---")

    rows = []
    for spl_folder, spl_short in SPL_MAPPING.items():
        per_shard_path = cases_dir / spl_folder / f"RQ2_perShard_{spl_folder}.xlsx"
        if not per_shard_path.exists():
            print(f"  [SKIP] {spl_short}: {per_shard_path.name} not found")
            continue
        for sheet_name, approach_label in approach_config.items():
            totals = aggregate_one_cell(per_shard_path, sheet_name)
            if totals is None:
                print(f"  [SKIP] {spl_short} {sheet_name}")
                continue
            row = build_breakdown_row(spl_short, sheet_name, approach_label, totals)
            if row is None:
                continue
            rows.append(row)
            print(f"  [OK]  {spl_short:5s} {sheet_name:15s}  "
                  f"T_pipe={row['T_pipeline(hours)']:>7.2f}h  "
                  f"SAT={row['SATShare(%)']:5.1f}%  "
                  f"Own={row['OwnCostShare(%)']:5.1f}%")

    if not rows:
        print(f"  ERROR: no data aggregated for {level}.")
        return

    df = pd.DataFrame(rows)
    df_display = df.drop(columns=["_phases", "_other"])

    # SPL row order
    df_display["_spl_order"] = df_display["SPL"].map(
        {s: i for i, s in enumerate(SPL_ORDER)})
    approach_order = {a: i for i, a in enumerate(approach_config.values())}
    df_display["_app_order"] = df_display["Approach"].map(approach_order)
    df_display = (df_display.sort_values(["_spl_order", "_app_order"])
                            .drop(columns=["_spl_order", "_app_order"])
                            .reset_index(drop=True))

    # Pivot tables
    sat_pivot = df_display.pivot_table(
        index="SPL", columns="Approach", values="SATShare(%)", aggfunc="first")
    sat_pivot = sat_pivot.reindex([s for s in SPL_ORDER if s in sat_pivot.index])
    sat_pivot = sat_pivot[[c for c in approach_config.values() if c in sat_pivot.columns]]

    own_pivot = df_display.pivot_table(
        index="SPL", columns="Approach", values="OwnCostShare(%)", aggfunc="first")
    own_pivot = own_pivot.reindex([s for s in SPL_ORDER if s in own_pivot.index])
    own_pivot = own_pivot[[c for c in approach_config.values() if c in own_pivot.columns]]

    # Output filenames carry the level suffix when not L2 so existing L2
    # outputs (manuscript main) are not overwritten.
    suffix = "" if level == "L2" else f"_{level}"
    out_xlsx = out_dir / f"rq2_time_breakdown{suffix}.xlsx"
    pdf_path = plots_dir / f"rq2_time_breakdown{suffix}.pdf"

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df_display.to_excel(writer, sheet_name="breakdown", index=False)
        sat_pivot.to_excel(writer, sheet_name="sat_share_pivot")
        own_pivot.to_excel(writer, sheet_name="own_cost_share_pivot")
    auto_fit_workbook(out_xlsx)
    print(f"  Saved: {out_xlsx.relative_to(project_root)}")

    # Stacked-bar PDF
    print(f"  Generating stacked-bar PDF ({level}) ...")
    plot_df = pd.DataFrame(rows)
    plot_df["_spl_order"] = plot_df["SPL"].map({s: i for i, s in enumerate(SPL_ORDER)})
    plot_df["_app_order"] = plot_df["Approach"].map(approach_order)
    plot_df = plot_df.sort_values(["_spl_order", "_app_order"]).reset_index(drop=True)

    fig, ax = plt.subplots(figsize=(max(12, 0.55 * len(plot_df)), 7))
    n = len(plot_df)
    x = np.arange(n)
    bottoms = np.zeros(n)

    for phase in COARSE_PHASES:
        if phase == "Other":
            heights = np.array([r["_other"] for r in plot_df.to_dict("records")]) / 1000 / 3600
        else:
            heights = np.array([r["_phases"][phase] for r in plot_df.to_dict("records")]) / 1000 / 3600
        if heights.sum() > 0:
            ax.bar(x, heights, bottom=bottoms, color=PHASE_COLORS[phase],
                   label=phase, edgecolor="white", linewidth=0.4)
            bottoms += heights

    x_labels = [f"{r['SPL']}\n{r['Approach'].split(' (')[0]}"
                for _, r in plot_df.iterrows()]
    ax.set_xticks(x)
    ax.set_xticklabels(x_labels, rotation=45, ha="right", fontsize=9)
    ax.set_ylabel("Cumulative SERIAL CPU time (hours)\n[80-shard sum of per-shard medians]", fontsize=10)
    ax.set_yscale("log")
    ax.set_title(f"RQ2: End-to-end pipeline time breakdown ({level})\n"
                 "(T_pipeline excludes Java's double-counted Transformation and validation-only coverage analysis)",
                 fontsize=11)
    ax.legend(loc="upper left", fontsize=9, framealpha=0.95)
    ax.grid(True, axis="y", alpha=0.3, which="both")
    plt.tight_layout()
    fig.savefig(pdf_path, format="pdf", bbox_inches="tight")
    plt.close(fig)
    print(f"  Saved: {pdf_path.relative_to(project_root)}")

    # Console summary
    print(f"\n  --- SATShare(%) by SPL x approach ({level}) ---")
    print("  " + sat_pivot.round(1).to_string().replace("\n", "\n  "))
    print(f"\n  --- OwnCostShare(%) by SPL x approach ({level}) ---")
    print("  " + own_pivot.round(1).to_string().replace("\n", "\n  "))


if __name__ == "__main__":
    main()