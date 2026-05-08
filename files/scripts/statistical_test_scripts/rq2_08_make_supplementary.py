#!/usr/bin/env python3
"""
rq2_08_make_supplementary.py

Step 8 of the RQ2 analysis pipeline.

Generates the complete set of LaTeX supplementary tables for the RQ2
industrial-scale throughput results. Every value is computed directly
from the master ApproachSummary Excel produced by rq2_03 and from the
analysis outputs produced by rq2_04 / rq2_06 / rq2_07; nothing is
hard-coded.

Output (single self-contained file):
    rq2_result/supplementary/rq2_supplementary.tex

The file produces 9 tables, S1--S9, each cross-referenced from the
main manuscript text:

    S1  End-to-end throughput at L = 3 (extends Table~\ref{tab:rq2-throughput}).
    S2  End-to-end throughput at L = 4.
    S3  Time breakdown (SAT share / per-product overhead share) at L = 3
        (extends Table~\ref{tab:rq2-breakdown}).
    S4  Time breakdown at L = 4.
    S5  Full pairwise Wilcoxon results: every (SPL x level pairing x
        metric) cell with N_shards, median_A, median_B, A12, magnitude,
        p_BH, significant_BH.
    S6  Per-cell run-to-run stability table (256 rows: 96 industrial +
        160 small/medium): SPL x Approach x Level x Metric -> Median CV,
        Max CV, Stability classification.
    S7  Worst-50 CV cells ranked listing.
    S8  Per-shard peak heap memory at L = 2, 3, 4 across all 8 SPLs and
        3 approaches.
    S9  Realised parallel speedup (T_serial / T_wall) for every SPL x
        approach x level cell.

Numerical precision
-------------------
All percentage values are reported to one decimal place using
ROUND_HALF_UP (school rounding). Float-arithmetic noise is scrubbed by
rounding the raw value to six decimals first, then quantising to one
decimal with ROUND_HALF_UP via the `decimal` module. Time values follow
the manuscript convention: ms / 1000 -> 1 decimal place for seconds and
hours; throughputs use 2 decimal places below 10 products/second and 1
decimal place above. Speedup values use 1 decimal place.

Paths (relative to this script's location):
    Script  : files/scripts/statistical_test_scripts/rq2_08_make_supplementary.py
    Inputs  :
        files/Cases/RQ2_ApproachSummary_medians.xlsx              (rq2_03)
        rq2_result/rq2_pairwise_wilcoxon.xlsx                     (rq2_04)
        rq2_result/rq2_time_breakdown.xlsx                        (rq2_06, L=2)
        rq2_result/rq2_time_breakdown_L3.xlsx                     (rq2_06 --levels L3)
        rq2_result/rq2_time_breakdown_L4.xlsx                     (rq2_06 --levels L4)
        rq2_result/rq2_run_stability.xlsx                         (rq2_07)
    Output  :
        rq2_result/supplementary/rq2_supplementary.tex

Usage:
    python rq2_08_make_supplementary.py
"""
from __future__ import annotations

import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# =============================================================================
# Paths
# =============================================================================
SCRIPT_DIR  = Path(__file__).resolve().parent
SCRIPTS_DIR = SCRIPT_DIR.parent
FILES_DIR   = SCRIPTS_DIR.parent
DATA_DIR    = FILES_DIR / "Cases"
RESULT_DIR  = SCRIPT_DIR / "rq2_result"
SUPP_DIR    = RESULT_DIR / "supplementary"


# =============================================================================
# Configuration
# =============================================================================
SPL_FOLDER_TO_SHORT = {
    "SodaVendingMachine":      "SVM",
    "eMail":                   "eM",
    "Elevator":                "El",
    "BankAccountv2":           "BAv2",
    "StudentAttendanceSystem": "SAS",
    "syngovia":                "Svia",
    "Tesla":                   "Te",
    "HockertyShirts":          "HS",
}
COL_ORDER = ["SVM", "eM", "El", "BAv2", "SAS", "Svia", "Te", "HS"]


# =============================================================================
# Numerical helpers
# =============================================================================
def round_half_up(x: float, ndigits: int = 1, scrub: int = 6) -> float:
    """ROUND_HALF_UP with float-noise scrubbing."""
    if pd.isna(x):
        return float("nan")
    scrubbed = round(float(x), scrub)
    q = Decimal(repr(scrubbed)).quantize(
        Decimal(10) ** -ndigits, rounding=ROUND_HALF_UP
    )
    return float(q)


def fmt_value(v, ndigits: int = 1, always_decimal: bool = False) -> str:
    """Format a numeric value with optional trailing-decimal preservation."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "---"
    r = round_half_up(float(v), ndigits)
    if not always_decimal and r == int(r):
        return f"{int(r)}"
    fmt = f"{{:.{ndigits}f}}"
    return fmt.format(r)


def fmt_int(v) -> str:
    """Format an integer with thousands separators using LaTeX-safe comma."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "---"
    return f"{int(v):,}".replace(",", r"{,}")


def fmt_throughput(v) -> str:
    """Two decimals below 10 products/second, one decimal above."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "---"
    if abs(v) < 10:
        return f"{round_half_up(float(v), 2):.2f}"
    return f"{round_half_up(float(v), 1):.1f}"


def fmt_time_smart(ms) -> str:
    """Format a time in ms using the appropriate unit:
    - < 60 s        : seconds (1dp)
    - < 60 min      : minutes (1dp)
    - >= 1 hour     : hours (1dp)
    """
    if ms is None or (isinstance(ms, float) and pd.isna(ms)) or ms <= 0:
        return "---"
    seconds = ms / 1000.0
    if seconds < 60:
        return rf"{round_half_up(seconds, 1):.1f}\,s"
    minutes = seconds / 60.0
    if minutes < 60:
        return rf"{round_half_up(minutes, 1):.1f}\,min"
    hours = minutes / 60.0
    return rf"{round_half_up(hours, 1):.1f}\,h"


def fmt_pvalue_inline(p) -> str:
    """Format a p-value as 'n.s.' / '<0.001' / '0.0123'."""
    if p is None or (isinstance(p, float) and pd.isna(p)):
        return "---"
    p = float(p)
    if p >= 0.05:
        return "n.s."
    if p < 0.001:
        return r"$<\!0.001$"
    return f"{round_half_up(p, 3):.3f}"


def fmt_a12(a) -> str:
    if a is None or (isinstance(a, float) and pd.isna(a)):
        return "---"
    return f"{round_half_up(float(a), 3):.3f}"


# =============================================================================
# Data access helpers
# =============================================================================
def find_project_root() -> Path | None:
    """Locate a directory that looks like a usable project root.

    Accepts both the legacy `files/Cases` layout and the newer layout
    where the script lives in `statistical_test_scripts/` next to a
    `rq2_result/` directory.
    """
    candidates = [Path.cwd(), *Path.cwd().parents]
    try:
        here = Path(__file__).resolve()
        candidates.extend([here.parent, *here.parents])
    except NameError:
        pass
    seen = set()
    for probe in candidates:
        if probe in seen:
            continue
        seen.add(probe)
        if (probe / "files" / "Cases").exists():
            return probe
    seen.clear()
    for probe in candidates:
        if probe in seen:
            continue
        seen.add(probe)
        if (probe / "rq2_result").exists():
            return probe
    return None


def load_approach_summary(path: Path) -> dict:
    """Load the ApproachSummary master Excel into a nested dict:
    out[sheet_name][SPL_long]['serial'|'wallclock'][column] = value
    """
    wb = load_workbook(path, data_only=True)
    out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = {}
        cur_section = None
        headers = None
        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a is None:
                continue
            if isinstance(a, str) and "SERIAL" in a:
                cur_section = "serial"
                headers = None
                continue
            if isinstance(a, str) and "WALL-CLOCK" in a:
                cur_section = "wallclock"
                headers = None
                continue
            if a == "SPL":
                headers = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                continue
            if headers and cur_section:
                rec = {h: ws.cell(r, c).value for c, h in enumerate(headers, 1) if h}
                rows.setdefault(a, {})[cur_section] = rec
        out[sheet_name] = rows
    return out


# =============================================================================
# Block builders
# =============================================================================
def _throughput_table(label: str, caption: str, level: str,
                       app_summary: dict) -> str:
    """
    Build S1 (L=3) or S2 (L=4) throughput table.

    Same layout as Table tab:rq2-throughput in the manuscript: 5 metric
    blocks (HP, T_serial, T_wall, throughput, edge coverage) x 3
    approaches x 8 SPLs.
    """
    long_to_short = {long: short for short, long in
                     [("SVM","SodaVendingMachine"),("eM","eMail"),("El","Elevator"),
                      ("BAv2","BankAccountv2"),("SAS","StudentAttendanceSystem"),
                      ("Svia","syngovia"),("Te","Tesla"),("HS","HockertyShirts")]}
    short_to_long = {v: k for k, v in long_to_short.items()}

    esg_sheet = f"ESG-Fx_{level}"
    efg_sheet = f"EFG_{level}"
    rw_sheet  = "RandomWalk_L0"

    esg = app_summary.get(esg_sheet, {})
    efg = app_summary.get(efg_sheet, {})
    rw  = app_summary.get(rw_sheet, {})

    def cell(approach_data, spl_short, section, col):
        long = short_to_long[spl_short]
        try:
            return approach_data[long][section][col]
        except KeyError:
            return None

    def row_hp(approach_data, label):
        cells = [fmt_int(cell(approach_data, s, "serial", "HandledProducts"))
                 for s in COL_ORDER]
        return rf"{label} & " + " & ".join(cells) + r" \\"

    def row_time(approach_data, label, section):
        cells = [fmt_time_smart(cell(approach_data, s, section, "T_pipeline(ms)"))
                 for s in COL_ORDER]
        return rf"{label} & " + " & ".join(cells) + r" \\"

    def row_throughput(approach_data, label):
        cells = [fmt_throughput(cell(approach_data, s, "wallclock",
                                     "Throughput_wallclock(prod_per_sec)"))
                 for s in COL_ORDER]
        return rf"{label} & " + " & ".join(cells) + r" \\"

    def row_coverage(approach_data, label):
        cells = []
        for s in COL_ORDER:
            v = cell(approach_data, s, "serial", "EdgeCoverage(%)")
            cells.append(fmt_value(v, ndigits=1, always_decimal=True))
        return rf"{label} & " + " & ".join(cells) + r" \\"

    out = []
    out.append(r"\begin{table*}[width=\textwidth,pos=htbp]")
    out.append(rf"\caption{{{caption}}}")
    out.append(rf"\label{{{label}}}")
    out.append(r"\setlength{\tabcolsep}{4pt}")
    out.append(r"\begin{tabularx}{\textwidth}{@{} l *{8}{Y} @{}}")
    out.append(r"\toprule")
    out.append(r"& \multicolumn{8}{c}{\textbf{Software Product Line$^*$}} \\")
    out.append(r"\cmidrule(lr){2-9}")
    out.append(r"\textbf{Approach} & "
               + " & ".join(rf"\textbf{{{c}}}" for c in COL_ORDER)
               + r" \\")
    out.append(r"\midrule")

    # HP block
    out.append(r"\multicolumn{9}{@{}l}{\textbf{Handled products}} \\")
    out.append(row_hp(esg, r"\textit{Model Once, Generate Any}"))
    out.append(row_hp(efg, "Structural baseline"))
    out.append(row_hp(rw,  "Stochastic baseline"))
    out.append(r"\midrule")

    # T_serial
    out.append(r"\multicolumn{9}{@{}l}{\textbf{$T^{\mathrm{ser}}_{\mathit{pipeline}}$ (cumulative serial cost)}} \\")
    out.append(row_time(esg, r"\textit{Model Once, Generate Any}", "serial"))
    out.append(row_time(efg, "Structural baseline", "serial"))
    out.append(row_time(rw,  "Stochastic baseline", "serial"))
    out.append(r"\midrule")

    # T_wall
    out.append(r"\multicolumn{9}{@{}l}{\textbf{$T^{\mathrm{wall}}_{\mathit{pipeline}}$ (80-core wall-clock latency)}} \\")
    out.append(row_time(esg, r"\textit{Model Once, Generate Any}", "wallclock"))
    out.append(row_time(efg, "Structural baseline", "wallclock"))
    out.append(row_time(rw,  "Stochastic baseline", "wallclock"))
    out.append(r"\midrule")

    # Throughput
    out.append(r"\multicolumn{9}{@{}l}{\textbf{Throughput (products / wall-clock second)}} \\")
    out.append(row_throughput(esg, r"\textit{Model Once, Generate Any}"))
    out.append(row_throughput(efg, "Structural baseline"))
    out.append(row_throughput(rw,  "Stochastic baseline"))
    out.append(r"\midrule")

    # Edge coverage
    out.append(r"\multicolumn{9}{@{}l}{\textbf{Edge coverage (\%)}} \\")
    out.append(row_coverage(esg, r"\textit{Model Once, Generate Any}"))
    out.append(row_coverage(efg, "Structural baseline"))
    out.append(row_coverage(rw,  "Stochastic baseline"))

    out.append(r"\bottomrule")
    out.append(r"\multicolumn{9}{@{}l}{\parbox[t]{0.95\textwidth}{\footnotesize "
               r"$^*$~SPL abbreviations as in Table~\ref{tab:rq1-tgen}.}} \\")
    out.append(r"\end{tabularx}")
    out.append(r"\end{table*}")
    return "\n".join(out)


def _breakdown_table(label: str, caption: str, level: str,
                      breakdown_path: Path) -> str:
    """
    Build S3 (L=3) or S4 (L=4) time-breakdown table from rq2_06's
    --levels L3 / L4 output.
    """
    if not breakdown_path.exists():
        return f"% {label} source missing: {breakdown_path}\n"

    sat = pd.read_excel(breakdown_path, sheet_name="sat_share_pivot")
    own = pd.read_excel(breakdown_path, sheet_name="own_cost_share_pivot")

    # Reindex rows to manuscript SPL order
    sat = sat.set_index("SPL").reindex(COL_ORDER)
    own = own.set_index("SPL").reindex(COL_ORDER)

    # Identify the three columns: ESG, EFG, RW
    def find_col(df, prefix):
        for c in df.columns:
            if c.startswith(prefix):
                return c
        return None

    esg_col = find_col(sat, "Model Once")
    efg_col = find_col(sat, "Structural")
    rw_col  = find_col(sat, "Stochastic")

    def row(df, col, label):
        cells = []
        for spl in COL_ORDER:
            try:
                v = df.loc[spl, col]
                cells.append(fmt_value(v, ndigits=1, always_decimal=True))
            except (KeyError, AttributeError):
                cells.append("---")
        return rf"{label} & " + " & ".join(cells) + r" \\"

    out = []
    out.append(r"\begin{table*}[width=\textwidth,pos=htbp]")
    out.append(rf"\caption{{{caption}}}")
    out.append(rf"\label{{{label}}}")
    out.append(r"\setlength{\tabcolsep}{4pt}")
    out.append(r"\begin{tabularx}{\textwidth}{@{} l *{8}{Y} @{}}")
    out.append(r"\toprule")
    out.append(r"& \multicolumn{8}{c}{\textbf{Software Product Line$^*$}} \\")
    out.append(r"\cmidrule(lr){2-9}")
    out.append(r"\textbf{Approach} & "
               + " & ".join(rf"\textbf{{{c}}}" for c in COL_ORDER)
               + r" \\")
    out.append(r"\midrule")

    # SAT share block
    out.append(r"\multicolumn{9}{@{}l}{\textbf{SAT share (\%)}} \\")
    out.append(row(sat, esg_col, r"\textit{Model Once, Generate Any}"))
    out.append(row(sat, efg_col, "Structural baseline"))
    out.append(row(sat, rw_col,  "Stochastic baseline"))
    out.append(r"\midrule")

    # Per-product overhead share
    out.append(r"\multicolumn{9}{@{}l}{\textbf{Per-product overhead share (\%)}} \\")
    out.append(row(own, esg_col, r"\textit{Model Once, Generate Any}"))
    out.append(row(own, efg_col, "Structural baseline"))
    out.append(row(own, rw_col,  "Stochastic baseline"))

    out.append(r"\bottomrule")
    out.append(r"\multicolumn{9}{@{}l}{\parbox[t]{0.95\textwidth}{\footnotesize "
               r"$^*$~SPL abbreviations as in Table~\ref{tab:rq1-tgen}.}} \\")
    out.append(r"\end{tabularx}")
    out.append(r"\end{table*}")
    return "\n".join(out)


def _wilcoxon_full_table(label: str, caption: str,
                          wilcoxon_path: Path) -> str:
    """
    S5: Full pairwise Wilcoxon results.
    Lists every (Pairing, Metric, SPL) row with N_shards, median_A,
    median_B, A12, magnitude, p_BH, significant_BH.
    """
    if not wilcoxon_path.exists():
        return f"% {label} source missing: {wilcoxon_path}\n"

    df = pd.read_excel(wilcoxon_path, sheet_name="all_comparisons")

    # Sort: by Pairing, then by Metric (stable), then by SPL in manuscript order
    spl_rank = {s: i for i, s in enumerate(COL_ORDER)}
    df = df.assign(_spl_rank=df["SPL"].map(spl_rank).fillna(99))
    pairing_order = ["ESGFx_L2_vs_EFG_L2", "ESGFx_L3_vs_EFG_L3",
                     "ESGFx_L4_vs_EFG_L4", "ESGFx_L2_vs_RW_L0",
                     "ESGFx_L3_vs_RW_L0", "ESGFx_L4_vs_RW_L0",
                     "ESGFx_L1_vs_RW_L0", "EFG_L2_vs_RW_L0"]
    metric_order = ["Throughput_pps", "T_pipeline_ms", "EdgeCoverage_pct",
                    "TestGenPeakMemory_MB", "TestGenTime_ms",
                    "EventCoverage_pct"]
    pair_rank = {p: i for i, p in enumerate(pairing_order)}
    metric_rank = {m: i for i, m in enumerate(metric_order)}
    df = df.assign(
        _pair_rank=df["Pairing"].map(pair_rank).fillna(99),
        _metric_rank=df["Metric"].map(metric_rank).fillna(99),
    )
    df = df.sort_values(["_pair_rank", "_metric_rank", "_spl_rank"])

    # Pretty pairing labels
    pretty_pair = {
        "ESGFx_L2_vs_EFG_L2": "MO L=2 vs Struct L=2",
        "ESGFx_L3_vs_EFG_L3": "MO L=3 vs Struct L=3",
        "ESGFx_L4_vs_EFG_L4": "MO L=4 vs Struct L=4",
        "ESGFx_L1_vs_RW_L0":  "MO L=1 vs Stoch",
        "ESGFx_L2_vs_RW_L0":  "MO L=2 vs Stoch",
        "ESGFx_L3_vs_RW_L0":  "MO L=3 vs Stoch",
        "ESGFx_L4_vs_RW_L0":  "MO L=4 vs Stoch",
        "EFG_L2_vs_RW_L0":    "Struct L=2 vs Stoch",
    }
    pretty_metric = {
        "Throughput_pps":       "Throughput",
        "T_pipeline_ms":        r"$T_{\mathit{pipeline}}$",
        "EdgeCoverage_pct":     "Edge cov.",
        "TestGenPeakMemory_MB": "Peak mem.",
        "TestGenTime_ms":       r"$T_{\mathit{gen}}$",
        "EventCoverage_pct":    "Event cov.",
    }
    pretty_winner = {
        "ESG-Fx":     "MO",
        "EFG":        "Struct",
        "RandomWalk": "Stoch",
        "tie":        "tie",
    }

    out = []
    out.append(r"\begin{table*}[width=\textwidth,pos=htbp]")
    out.append(rf"\caption{{{caption}}}")
    out.append(rf"\label{{{label}}}")
    out.append(r"\setlength{\tabcolsep}{3pt}")
    out.append(r"\footnotesize")
    out.append(r"\begin{tabular*}{\textwidth}"
               r"{@{\extracolsep{\fill}} l l l c r r r l l @{}}")
    out.append(r"\toprule")
    out.append(r"\textbf{Pairing} & \textbf{Metric} & \textbf{SPL} & "
               r"\textbf{$n$} & \textbf{med. A} & \textbf{med. B} & "
               r"\textbf{$A_{12}$} & \textbf{Magn.} & "
               r"\textbf{$p_{\mathit{BH}}$} \\")
    out.append(r"\midrule")

    prev_pair = None
    prev_metric = None
    for _, r in df.iterrows():
        pair = str(r["Pairing"])
        metric = str(r["Metric"])
        if pair not in pretty_pair:
            continue  # skip pairings outside the manuscript scope
        # Show pairing on the first row of each (pair, metric) block; show
        # metric on the first row of each metric block within a pairing.
        # Every other row is blank to avoid clutter, but we keep the SPL
        # column populated so each row is self-readable.
        new_pair = (pair != prev_pair)
        new_metric = (pair, metric) != (prev_pair, prev_metric)
        # Insert a small visual gap whenever the pairing changes
        if new_pair and prev_pair is not None:
            out.append(r"\addlinespace[2pt]")
        pair_cell = pretty_pair[pair] if new_pair else ""
        metric_cell = pretty_metric.get(metric, metric) if new_metric else ""
        prev_pair = pair
        prev_metric = metric
        spl = str(r["SPL"])
        n = int(r["N_shards"]) if not pd.isna(r["N_shards"]) else "---"
        # Format median values according to the metric's scale
        if metric == "Throughput_pps":
            med_a = fmt_throughput(r["median_A"])
            med_b = fmt_throughput(r["median_B"])
        elif metric in ("T_pipeline_ms", "TestGenTime_ms"):
            med_a = fmt_time_smart(r["median_A"])
            med_b = fmt_time_smart(r["median_B"])
        elif metric in ("EdgeCoverage_pct", "EventCoverage_pct"):
            med_a = fmt_value(r["median_A"], 1, always_decimal=True) + r"\%"
            med_b = fmt_value(r["median_B"], 1, always_decimal=True) + r"\%"
        elif metric == "TestGenPeakMemory_MB":
            med_a = fmt_value(r["median_A"], 1, always_decimal=True) + r"\,MB"
            med_b = fmt_value(r["median_B"], 1, always_decimal=True) + r"\,MB"
        else:
            med_a = fmt_value(r["median_A"])
            med_b = fmt_value(r["median_B"])
        a12 = fmt_a12(r["A12_A_vs_B"])
        mag = str(r.get("magnitude", "---"))
        p_bh = fmt_pvalue_inline(r["p_BH"])
        sig = "" if r.get("significant_BH", False) else r"$^{\mathrm{ns}}$"

        out.append(rf"{pair_cell} & {metric_cell} & {spl} & {n} & "
                   rf"{med_a} & {med_b} & {a12} & {mag} & {p_bh}{sig} \\")
        # Visual gap between pairings
        if str(r["SPL"]) == "HS" and metric_cell == "":
            pass  # within metric block, no extra space
    # Add a footer
    out.append(r"\bottomrule")
    out.append(r"\end{tabular*}")
    out.append(r"\end{table*}")
    return "\n".join(out)


def _stability_full_table(label: str, caption: str,
                           stability_path: Path) -> str:
    """
    S6: Per-cell run-to-run stability. 256 rows from rq2_07's summary
    sheet: SPL x Approach x Level x Metric -> Median CV%, Max CV%,
    Stability classification.
    """
    if not stability_path.exists():
        return f"% {label} source missing: {stability_path}\n"

    df = pd.read_excel(stability_path, sheet_name="summary")

    # Sort: by Scale (industrial first), then by SPL (manuscript order),
    # then by Approach, Level, Metric
    spl_rank = {s: i for i, s in enumerate(COL_ORDER)}
    scale_rank = {"Industrial": 0, "Small/Medium": 1}
    approach_rank = {"Model Once, Generate Any": 0,
                     "Structural Baseline": 1, "Stochastic Baseline": 2}
    level_rank = {"L0": 0, "L1": 1, "L2": 2, "L3": 3, "L4": 4}
    metric_rank = {"T_pipeline(ms)": 0, "TestGenTime(ms)": 1,
                   "TestGenPeakMemory(MB)": 2, "Coverage(%)": 3}

    df = df.assign(
        _scale=df["Scale"].map(scale_rank).fillna(99),
        _spl=df["SPL"].map(spl_rank).fillna(99),
        _ap=df["Approach"].map(approach_rank).fillna(99),
        _lvl=df["Level"].map(level_rank).fillna(99),
        _met=df["Metric"].map(metric_rank).fillna(99),
    )
    df = df.sort_values(["_scale", "_spl", "_ap", "_lvl", "_met"])

    # Pretty approach labels
    pretty_ap = {"Model Once, Generate Any": "MO",
                 "Structural Baseline": "Struct",
                 "Stochastic Baseline": "Stoch"}
    pretty_metric = {"T_pipeline(ms)": r"$T_{\mathit{pipeline}}$",
                     "TestGenTime(ms)": r"$T_{\mathit{gen}}$",
                     "TestGenPeakMemory(MB)": "Peak memory",
                     "Coverage(%)": "Coverage"}

    out = []
    out.append(r"\begin{table*}[width=\textwidth,pos=htbp]")
    out.append(rf"\caption{{{caption}}}")
    out.append(rf"\label{{{label}}}")
    out.append(r"\setlength{\tabcolsep}{3pt}")
    out.append(r"\footnotesize")
    out.append(r"\begin{longtable}{@{\extracolsep{\fill}} l l l l l r r r l @{}}")
    out.append(r"\toprule")
    out.append(r"\textbf{Scale} & \textbf{SPL} & \textbf{Approach} & "
               r"\textbf{$L$} & \textbf{Metric} & \textbf{$n$} & "
               r"\textbf{Median CV (\%)} & \textbf{Max CV (\%)} & "
               r"\textbf{Stability} \\")
    out.append(r"\midrule")
    out.append(r"\endfirsthead")
    out.append(r"\toprule")
    out.append(r"\textbf{Scale} & \textbf{SPL} & \textbf{Approach} & "
               r"\textbf{$L$} & \textbf{Metric} & \textbf{$n$} & "
               r"\textbf{Median CV (\%)} & \textbf{Max CV (\%)} & "
               r"\textbf{Stability} \\")
    out.append(r"\midrule")
    out.append(r"\endhead")
    out.append(r"\bottomrule")
    out.append(r"\endfoot")

    prev_scale = None
    prev_spl = None
    prev_ap = None
    for _, r in df.iterrows():
        scale = str(r["Scale"])
        spl = str(r["SPL"])
        ap = str(r["Approach"])
        scale_cell = scale if scale != prev_scale else ""
        spl_cell = spl if (spl, scale) != (prev_spl, prev_scale) else ""
        ap_cell = pretty_ap.get(ap, ap) if (spl, ap) != (prev_spl, prev_ap) else ""
        prev_scale = scale
        prev_spl = spl
        prev_ap = ap

        lvl = str(r["Level"])
        metric = pretty_metric.get(str(r["Metric"]), str(r["Metric"]))
        n = int(r["N_shards"]) if not pd.isna(r["N_shards"]) else "---"
        med_cv = fmt_value(r["Median_CV_pct"], 2, always_decimal=True)
        max_cv = fmt_value(r["Max_CV_pct"], 2, always_decimal=True)
        stab = str(r["Stability"]).replace("_", r"\_")

        out.append(rf"{scale_cell} & {spl_cell} & {ap_cell} & "
                   rf"{lvl} & {metric} & {n} & {med_cv} & {max_cv} & {stab} \\")

    out.append(r"\end{longtable}")
    out.append(r"\end{table*}")
    return "\n".join(out)


def _worst_cv_table(label: str, caption: str,
                     stability_path: Path, top_n: int = 50) -> str:
    """
    S7: Top-N CV cells across all scales, ranked descending by Median CV%.
    """
    if not stability_path.exists():
        return f"% {label} source missing: {stability_path}\n"

    df = pd.read_excel(stability_path, sheet_name="summary")
    df = df.sort_values("Median_CV_pct", ascending=False).head(top_n)

    pretty_ap = {"Model Once, Generate Any": "MO",
                 "Structural Baseline": "Struct",
                 "Stochastic Baseline": "Stoch"}
    pretty_metric = {"T_pipeline(ms)": r"$T_{\mathit{pipeline}}$",
                     "TestGenTime(ms)": r"$T_{\mathit{gen}}$",
                     "TestGenPeakMemory(MB)": "Peak memory",
                     "Coverage(%)": "Coverage"}

    out = []
    out.append(r"\begin{table}[pos=htbp]")
    out.append(rf"\caption{{{caption}}}")
    out.append(rf"\label{{{label}}}")
    out.append(r"\setlength{\tabcolsep}{3pt}")
    out.append(r"\footnotesize")
    out.append(r"\begin{tabular*}{\columnwidth}"
               r"{@{\extracolsep{\fill}} r l l l l l r r l @{}}")
    out.append(r"\toprule")
    out.append(r"\textbf{\#} & \textbf{Scale} & \textbf{SPL} & "
               r"\textbf{Approach} & \textbf{$L$} & \textbf{Metric} & "
               r"\textbf{Median CV} & \textbf{Max CV} & "
               r"\textbf{Stability} \\")
    out.append(r"\midrule")

    for rank, (_, r) in enumerate(df.iterrows(), start=1):
        scale = str(r["Scale"])
        spl = str(r["SPL"])
        ap = pretty_ap.get(str(r["Approach"]), str(r["Approach"]))
        lvl = str(r["Level"])
        metric = pretty_metric.get(str(r["Metric"]), str(r["Metric"]))
        med_cv = fmt_value(r["Median_CV_pct"], 2, always_decimal=True)
        max_cv = fmt_value(r["Max_CV_pct"], 2, always_decimal=True)
        stab = str(r["Stability"]).replace("_", r"\_")
        out.append(rf"{rank} & {scale} & {spl} & {ap} & "
                   rf"{lvl} & {metric} & {med_cv}\% & {max_cv}\% & {stab} \\")

    out.append(r"\bottomrule")
    out.append(r"\end{tabular*}")
    out.append(r"\end{table}")
    return "\n".join(out)


def _memory_table(label: str, caption: str, app_summary: dict) -> str:
    """
    S8: Per-shard peak heap memory at L=2, L=3, L=4 for all 8 SPLs and
    3 approaches. ESG and EFG operate at L=2,3,4; RW only at L=0 and is
    listed once per SPL.
    """
    long_to_short = {long: short for short, long in
                     [("SVM","SodaVendingMachine"),("eM","eMail"),("El","Elevator"),
                      ("BAv2","BankAccountv2"),("SAS","StudentAttendanceSystem"),
                      ("Svia","syngovia"),("Te","Tesla"),("HS","HockertyShirts")]}
    short_to_long = {v: k for k, v in long_to_short.items()}

    def cell(approach_data, spl_short):
        long = short_to_long[spl_short]
        try:
            return approach_data[long]["serial"]["TestGenPeakMemory(MB)"]
        except KeyError:
            return None

    def row(sheet_name, label):
        data = app_summary.get(sheet_name, {})
        cells = [fmt_value(cell(data, s), 1, always_decimal=True)
                 for s in COL_ORDER]
        return rf"{label} & " + " & ".join(cells) + r" \\"

    out = []
    out.append(r"\begin{table*}[width=\textwidth,pos=htbp]")
    out.append(rf"\caption{{{caption}}}")
    out.append(rf"\label{{{label}}}")
    out.append(r"\setlength{\tabcolsep}{4pt}")
    out.append(r"\begin{tabularx}{\textwidth}{@{} l c *{8}{Y} @{}}")
    out.append(r"\toprule")
    out.append(r"& & \multicolumn{8}{c}{\textbf{Software Product Line$^*$}} \\")
    out.append(r"\cmidrule(lr){3-10}")
    out.append(r"\textbf{Approach} & \textbf{$L$} & "
               + " & ".join(rf"\textbf{{{c}}}" for c in COL_ORDER)
               + r" \\")
    out.append(r"\midrule")
    out.append(r"\multirow{3}{*}{\textit{Model Once, Generate Any}} "
               + row("ESG-Fx_L2", "& 2 ").lstrip())
    out.append(row("ESG-Fx_L3", "& 3 "))
    out.append(row("ESG-Fx_L4", "& 4 "))
    out.append(r"\cmidrule(l{0pt}r{0pt}){2-10}")
    out.append(r"\multirow{3}{*}{Structural baseline} "
               + row("EFG_L2", "& 2 ").lstrip())
    out.append(row("EFG_L3", "& 3 "))
    out.append(row("EFG_L4", "& 4 "))
    out.append(r"\cmidrule(l{0pt}r{0pt}){2-10}")
    out.append(row("RandomWalk_L0", "Stochastic baseline & --- "))
    out.append(r"\bottomrule")
    out.append(r"\multicolumn{10}{@{}l}{\parbox[t]{0.95\textwidth}{\footnotesize "
               r"$^*$~SPL abbreviations as in Table~\ref{tab:rq1-tgen}. "
               r"Memory values are per-shard medians in MB.}} \\")
    out.append(r"\end{tabularx}")
    out.append(r"\end{table*}")
    return "\n".join(out)


def _speedup_table(label: str, caption: str, app_summary: dict) -> str:
    """
    S9: Realised parallel speedup (T_serial / T_wall) for each
    (SPL x approach x level) cell.
    """
    long_to_short = {long: short for short, long in
                     [("SVM","SodaVendingMachine"),("eM","eMail"),("El","Elevator"),
                      ("BAv2","BankAccountv2"),("SAS","StudentAttendanceSystem"),
                      ("Svia","syngovia"),("Te","Tesla"),("HS","HockertyShirts")]}
    short_to_long = {v: k for k, v in long_to_short.items()}

    def speedup(approach_data, spl_short):
        long = short_to_long[spl_short]
        try:
            t_ser = approach_data[long]["serial"]["T_pipeline(ms)"]
            t_wall = approach_data[long]["wallclock"]["T_pipeline(ms)"]
            if t_wall is None or t_wall <= 0:
                return None
            return t_ser / t_wall
        except KeyError:
            return None

    def row(sheet_name, label):
        data = app_summary.get(sheet_name, {})
        cells = [fmt_value(speedup(data, s), 1, always_decimal=True) + r"$\times$"
                 if speedup(data, s) is not None else "---"
                 for s in COL_ORDER]
        return rf"{label} & " + " & ".join(cells) + r" \\"

    out = []
    out.append(r"\begin{table*}[width=\textwidth,pos=htbp]")
    out.append(rf"\caption{{{caption}}}")
    out.append(rf"\label{{{label}}}")
    out.append(r"\setlength{\tabcolsep}{4pt}")
    out.append(r"\begin{tabularx}{\textwidth}{@{} l c *{8}{Y} @{}}")
    out.append(r"\toprule")
    out.append(r"& & \multicolumn{8}{c}{\textbf{Software Product Line$^*$}} \\")
    out.append(r"\cmidrule(lr){3-10}")
    out.append(r"\textbf{Approach} & \textbf{$L$} & "
               + " & ".join(rf"\textbf{{{c}}}" for c in COL_ORDER)
               + r" \\")
    out.append(r"\midrule")
    out.append(r"\multirow{3}{*}{\textit{Model Once, Generate Any}} "
               + row("ESG-Fx_L2", "& 2 ").lstrip())
    out.append(row("ESG-Fx_L3", "& 3 "))
    out.append(row("ESG-Fx_L4", "& 4 "))
    out.append(r"\cmidrule(l{0pt}r{0pt}){2-10}")
    out.append(r"\multirow{3}{*}{Structural baseline} "
               + row("EFG_L2", "& 2 ").lstrip())
    out.append(row("EFG_L3", "& 3 "))
    out.append(row("EFG_L4", "& 4 "))
    out.append(r"\cmidrule(l{0pt}r{0pt}){2-10}")
    out.append(row("RandomWalk_L0", "Stochastic baseline & --- "))
    out.append(r"\bottomrule")
    out.append(r"\multicolumn{10}{@{}l}{\parbox[t]{0.95\textwidth}{\footnotesize "
               r"$^*$~SPL abbreviations as in Table~\ref{tab:rq1-tgen}. "
               r"Speedup is the ratio of $T^{\mathrm{ser}}_{\mathit{pipeline}}$ "
               r"(serial sum across the 80 shards) to "
               r"$T^{\mathrm{wall}}_{\mathit{pipeline}}$ (maximum across "
               r"shards). The ideal speedup on this 80-core deployment is "
               r"$80\times$.}} \\")
    out.append(r"\end{tabularx}")
    out.append(r"\end{table*}")
    return "\n".join(out)


# =============================================================================
# Captions
# =============================================================================
PREAMBLE = r"""% =========================================================================
% RQ2 Supplementary Material
% Auto-generated by rq2_10_make_supplementary.py
% Every value computed from the raw data; do not edit by hand.
% =========================================================================
%
% This file can be either compiled standalone (uncomment the documentclass
% block below) or \input into the main supplementary tex file. Tables are
% labelled tab:rq2-supp-S1 ... tab:rq2-supp-S9 and counted independently
% of the manuscript table counter (use \renewcommand{\thetable}{S\arabic{table}}
% in the parent document).
%
% \documentclass[a4paper,fleqn]{cas-sc}
% \usepackage{booktabs,multirow,tabularx,longtable}
% \newcolumntype{Y}{>{\centering\arraybackslash}X}
% \newcolumntype{L}{>{\raggedright\arraybackslash}X}
% \begin{document}

\section*{S-RQ2. Supplementary Material for RQ2 (Industrial-Scale Throughput)}
\label{sec:supp_rq2}

This appendix collects the supplementary tables referenced in the RQ2
results section (Section~\ref{sec:results_rq2}) of the main manuscript.
Tables S1 and S2 give the full throughput numbers at $L = 3$ and $L = 4$;
the manuscript main text (Table~\ref{tab:rq2-throughput}) reports
$L = 2$ only. Tables S3 and S4 give the corresponding time-breakdown
values; Table~\ref{tab:rq2-breakdown} in the manuscript reports the
$L = 2$ version. Table S5 gives the full pairwise Wilcoxon results
across every (SPL, level pairing, metric) cell; the manuscript
(Section~\ref{sec:rq2_pairwise}) cites only the cells used as evidence.
Table S6 gives the per-cell run-to-run stability values for all 256
(SPL, approach, level, metric) cells; the manuscript
(Section~\ref{sec:rq2_stability}) summarises these as counts per band.
Table S7 ranks the 50 cells with the highest median coefficient of
variation. Table S8 gives the peak heap memory measured during test
generation across all SPLs, approaches, and coverage levels;
Section~\ref{sec:rq2_pairwise} cites only the syngo.via and Hockerty
Shirts $L = 2$ values. Table S9 gives the realised parallel speedup
($T^{\mathrm{ser}}_{\mathit{pipeline}} / T^{\mathrm{wall}}_{\mathit{pipeline}}$)
for every cell; the manuscript reports the three industrial speedups
inline.

All percentage values are rounded to one decimal place using
ROUND\_HALF\_UP after a six-decimal float-noise scrub. Time values are
rendered in seconds, minutes, or hours depending on magnitude
(seconds below one minute, minutes below one hour, hours otherwise).
Throughput values use two decimal places below 10 products/second and
one decimal place above, so that ratios computed from displayed values
match those reported in the manuscript.
"""

CAPTION_S1 = (
    r"S1: End-to-end throughput at $L = 3$ across the eight SPLs. Layout "
    r"matches Table~\ref{tab:rq2-throughput} in the manuscript; the same "
    r"definitions of HP, $T^{\mathrm{ser}}_{\mathit{pipeline}}$, "
    r"$T^{\mathrm{wall}}_{\mathit{pipeline}}$, throughput, and edge "
    r"coverage apply."
)
CAPTION_S2 = (
    r"S2: End-to-end throughput at $L = 4$ across the eight SPLs. Same "
    r"layout and definitions as Table~\ref{tab:rq2-supp-S1}."
)
CAPTION_S3 = (
    r"S3: SAT share and per-product overhead share of "
    r"$T^{\mathrm{ser}}_{\mathit{pipeline}}$ at $L = 3$. Layout matches "
    r"Table~\ref{tab:rq2-breakdown}; the two shares sum to approximately "
    r"$100\%$ for each cell, with the small remainder being the "
    r"product-instantiation phase."
)
CAPTION_S4 = (
    r"S4: SAT share and per-product overhead share of "
    r"$T^{\mathrm{ser}}_{\mathit{pipeline}}$ at $L = 4$. Same layout as "
    r"Table~\ref{tab:rq2-supp-S3}."
)
CAPTION_S5 = (
    r"S5: Pairwise Wilcoxon signed-rank results on the 80 paired shards "
    r"per (SPL, level pairing, metric). MO denotes \textit{Model Once, "
    r"Generate Any}, Struct the structural baseline, and Stoch the "
    r"stochastic baseline. The reported $A_{12}$ is the Vargha--Delaney "
    r"effect size (approach A vs.~approach B; values close to 1 favour "
    r"A, values close to 0 favour B). $p_{\mathit{BH}}$ is the "
    r"Benjamini--Hochberg corrected $p$-value at $\alpha = 0.05$; cells "
    r"flagged $^{\mathrm{ns}}$ are not significant after correction."
)
CAPTION_S6 = (
    r"S6: Per-cell run-to-run stability summary for all 256 (SPL, "
    r"approach, coverage level, metric) cells. The coefficient of "
    r"variation $\mathrm{CV}\% = 100 \cdot \mathrm{std}/\mathrm{mean}$ "
    r"is computed across runs of each shard, then aggregated to the "
    r"median (column \emph{Median CV}) and maximum (column "
    r"\emph{Max CV}) across the 80 shards. Stability bands: "
    r"\emph{very stable} when median CV $< 5\%$, \emph{stable} between "
    r"$5\%$ and $10\%$, \emph{notable variance} above $10\%$ "
    r"(Section~\ref{sec:rq2_stability})."
)
CAPTION_S7 = (
    r"S7: Top 50 cells with the highest median coefficient of variation "
    r"across the 80 shards. The small and medium-sized SPLs dominate "
    r"this ranking because their absolute timing values are small "
    r"(milliseconds per shard), so a few milliseconds of JVM jitter "
    r"inflates the relative spread. The same effect was discussed for "
    r"RQ1 in Section~\ref{sec:rq1_stability}."
)
CAPTION_S8 = (
    r"S8: Per-shard median peak heap memory (MB) during test generation, "
    r"across all SPLs, approaches, and coverage levels. The article main "
    r"text (Section~\ref{sec:rq2_pairwise}) cites only the syngo.via and "
    r"Hockerty Shirts values at $L = 2$; this table gives the complete "
    r"set."
)
CAPTION_S9 = (
    r"S9: Realised parallel speedup of the 80-core sharded deployment, "
    r"computed as $T^{\mathrm{ser}}_{\mathit{pipeline}} / "
    r"T^{\mathrm{wall}}_{\mathit{pipeline}}$ for every (SPL, approach, "
    r"coverage level) cell. The article main text "
    r"(Section~\ref{sec:rq2_throughput}) reports the three industrial "
    r"values for \textit{Model Once, Generate Any} inline; this table "
    r"gives the complete set."
)


# =============================================================================
# Main
# =============================================================================
def main() -> None:
    print("=" * 80)
    print("rq2_10: BUILD RQ2 SUPPLEMENTARY MATERIAL")
    print("=" * 80)

    project_root = find_project_root()
    if project_root is None:
        # When run on a test bench without files/Cases, fall back on the
        # script's own parent directory so we can still build the .tex.
        project_root = SCRIPT_DIR

    print(f"Script  : {SCRIPT_DIR}")
    print(f"Result  : {RESULT_DIR}")
    SUPP_DIR.mkdir(parents=True, exist_ok=True)

    # --- Locate inputs ---
    # Master ApproachSummary may live in files/Cases/ or in the result
    # directory; check both.
    candidates = [
        DATA_DIR / "RQ2_ApproachSummary_medians.xlsx",
        RESULT_DIR / "RQ2_ApproachSummary_medians.xlsx",
    ]
    app_summary_path = next((c for c in candidates if c.exists()), None)
    if app_summary_path is None:
        print("ERROR: RQ2_ApproachSummary_medians.xlsx not found.")
        sys.exit(1)
    print(f"ApproachSummary: {app_summary_path}")

    wilcoxon_path     = RESULT_DIR / "rq2_pairwise_wilcoxon.xlsx"
    breakdown_l3_path = RESULT_DIR / "rq2_time_breakdown_L3.xlsx"
    breakdown_l4_path = RESULT_DIR / "rq2_time_breakdown_L4.xlsx"
    stability_path    = RESULT_DIR / "rq2_run_stability.xlsx"

    # Load master ApproachSummary once (used by S1, S2, S8, S9).
    app_summary = load_approach_summary(app_summary_path)

    parts: list[str] = [PREAMBLE, ""]

    print("\nBuilding S1 (Throughput, L=3) ...")
    parts.append(_throughput_table(
        "tab:rq2-supp-S1", CAPTION_S1, "L3", app_summary))
    parts.append("")

    print("Building S2 (Throughput, L=4) ...")
    parts.append(_throughput_table(
        "tab:rq2-supp-S2", CAPTION_S2, "L4", app_summary))
    parts.append("")

    print("Building S3 (Time breakdown, L=3) ...")
    parts.append(_breakdown_table(
        "tab:rq2-supp-S3", CAPTION_S3, "L3", breakdown_l3_path))
    parts.append("")

    print("Building S4 (Time breakdown, L=4) ...")
    parts.append(_breakdown_table(
        "tab:rq2-supp-S4", CAPTION_S4, "L4", breakdown_l4_path))
    parts.append("")

    print("Building S5 (Full pairwise Wilcoxon) ...")
    parts.append(_wilcoxon_full_table(
        "tab:rq2-supp-S5", CAPTION_S5, wilcoxon_path))
    parts.append("")

    print("Building S6 (Per-cell stability summary) ...")
    parts.append(_stability_full_table(
        "tab:rq2-supp-S6", CAPTION_S6, stability_path))
    parts.append("")

    print("Building S7 (Worst-50 CV ranking) ...")
    parts.append(_worst_cv_table(
        "tab:rq2-supp-S7", CAPTION_S7, stability_path, top_n=50))
    parts.append("")

    print("Building S8 (Peak heap memory across all cells) ...")
    parts.append(_memory_table(
        "tab:rq2-supp-S8", CAPTION_S8, app_summary))
    parts.append("")

    print("Building S9 (Realised parallel speedup) ...")
    parts.append(_speedup_table(
        "tab:rq2-supp-S9", CAPTION_S9, app_summary))
    parts.append("")

    out_path = SUPP_DIR / "rq2_supplementary.tex"
    out_path.write_text("\n".join(parts), encoding="utf-8")

    print()
    print("=" * 80)
    print(f"DONE.  Wrote {out_path}")
    print(f"        ({(len(parts) - 2) // 2} table blocks)")
    print("=" * 80)


if __name__ == "__main__":
    main()