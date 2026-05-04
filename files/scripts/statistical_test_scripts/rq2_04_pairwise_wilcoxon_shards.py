#!/usr/bin/env python3
"""
rq2_04_pairwise_wilcoxon_shards.py  (RQ2 Step 4 -- reviewer armor)

Paired Wilcoxon signed-rank tests on the 80 shards within each SPL.
The shards form a true paired design: shard k of approach A processes
EXACTLY the same product partition as shard k of approach B (the
sharding rule (productID-1) mod 80 is approach-independent), so paired
comparison on per-shard medians is statistically valid.

What this test answers
----------------------
Per (SPL x level-pairing x metric):
    "On the 80 paired shards, did approach A use systematically less
     time / less memory / more coverage than approach B?"

What it does NOT answer
-----------------------
"Is the practical effect large?" -- N=80 paired observations means that
even a tiny systematic difference yields p ~ 1e-15. P-values are not
the point at this scale. The columns that carry the actual story are:

    median_delta_A_minus_B  : sign and magnitude of the typical gap
    A12_A_vs_B              : Vargha-Delaney effect size
                              (>0.71 = large; <0.5 means A < B)
    winner                  : interpretation given the metric direction
                              (lower-is-better for time/memory,
                               higher-is-better for coverage)

Metric definitions (canonical schema)
-------------------------------------
T_pipeline(ms)             : end-to-end pipeline (NOT Java's buggy total)
TestGenTime(ms)            : approach-specific test generation cost
TestGenPeakMemory(MB)      : peak heap during test generation
EdgeCoverage(%)            : edge coverage attained
                             (for ESG-Fx_L1 this column doesn't exist;
                              EventCoverage(%) is used instead -- the
                              comparison is then ESG-Fx_L1's event
                              coverage vs the same metric on EFG/RW)

Level pairings
--------------
ESG-Fx Lk vs EFG Lk for k in {2,3,4}
ESG-Fx_L1 vs RandomWalk_L0  (event-coverage pairing)
ESG-Fx Lk vs RandomWalk_L0 for k in {2,3,4}
EFG_L2    vs RandomWalk_L0

Output
------
files/scripts/statistical_test_scripts/rq2_result/rq2_pairwise_wilcoxon.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from scipy.stats import wilcoxon


# ─── Configuration ─────────────────────────────────────────────────────
SPL_MAPPING = {
    "SodaVendingMachine":      "SVM",
    "eMail":                   "eM",
    "Elevator":                "El",
    "BankAccountv2":           "BAv2",
    "StudentAttendanceSystem": "SAS",
    "syngovia":                "Svia",
    "Tesla":                   "Te",
    "HockertyShirts":          "HS",
}
SPL_ORDER = ["SVM", "eM", "El", "BAv2", "SAS", "Svia", "Te", "HS"]

APPROACH_LABEL = {
    "ESG-Fx":     "Model Once, Generate Any",
    "EFG":        "Structural Baseline",
    "RandomWalk": "Stochastic Baseline",
}

# Level pairings: (approachA, levelA, approachB, levelB, label)
LEVEL_PAIRINGS = [
    ("ESG-Fx", "L2", "EFG",        "L2", "ESGFx_L2_vs_EFG_L2"),
    ("ESG-Fx", "L3", "EFG",        "L3", "ESGFx_L3_vs_EFG_L3"),
    ("ESG-Fx", "L4", "EFG",        "L4", "ESGFx_L4_vs_EFG_L4"),
    ("ESG-Fx", "L1", "RandomWalk", "L0", "ESGFx_L1_vs_RW_L0"),
    ("ESG-Fx", "L2", "RandomWalk", "L0", "ESGFx_L2_vs_RW_L0"),
    ("ESG-Fx", "L3", "RandomWalk", "L0", "ESGFx_L3_vs_RW_L0"),
    ("ESG-Fx", "L4", "RandomWalk", "L0", "ESGFx_L4_vs_RW_L0"),
    ("EFG",    "L2", "RandomWalk", "L0", "EFG_L2_vs_RW_L0"),
]

# Metrics: (canonical_column, output_label, lower_is_better)
# For the coverage metric we route per-pairing -- see resolve_coverage_col.
# Throughput is derived per shard from HandledProducts and T_pipeline(ms);
# higher is better, opposite direction from time.
METRICS = [
    ("__throughput__",        "Throughput_pps",       False),
    ("T_pipeline(ms)",        "T_pipeline_ms",        True),
    ("TestGenTime(ms)",       "TestGenTime_ms",       True),
    ("TestGenPeakMemory(MB)", "TestGenPeakMemory_MB", True),
    # Coverage handled separately so we can pick Event vs Edge per pairing.
    ("__coverage__",          "Coverage_pct",         False),
]


def resolve_coverage_col(approachA, levelA, approachB, levelB):
    """ESG-Fx_L1 only has EventCoverage(%); ESG-Fx L>=2 only has
    EdgeCoverage(%); EFG and RW have both. For a fair pairwise
    comparison, pick the column that BOTH sides have.

    Returns (colA, colB, label) or (None, None, 'NA') when no common
    coverage column exists (shouldn't happen with our pairings, but
    we're defensive).
    """
    has_event = lambda a, l: (a == "ESG-Fx" and l == "L1") or a in ("EFG", "RandomWalk")
    has_edge  = lambda a, l: (a == "ESG-Fx" and l != "L1") or a in ("EFG", "RandomWalk")

    if has_edge(approachA, levelA) and has_edge(approachB, levelB):
        return "EdgeCoverage(%)", "EdgeCoverage(%)", "EdgeCoverage_pct"
    if has_event(approachA, levelA) and has_event(approachB, levelB):
        return "EventCoverage(%)", "EventCoverage(%)", "EventCoverage_pct"
    return None, None, "Coverage_NA"


# ─── Statistics ────────────────────────────────────────────────────────
def vargha_delaney_a12(x, y):
    x = np.asarray(x, dtype=float); y = np.asarray(y, dtype=float)
    if len(x) == 0 or len(y) == 0:
        return np.nan
    diff = x[:, None] - y[None, :]
    gt = int(np.sum(diff > 0))
    eq = int(np.sum(diff == 0))
    return (gt + 0.5 * eq) / (len(x) * len(y))


def a12_magnitude(a12):
    """Vargha-Delaney magnitude bins from Vargha & Delaney (2000)."""
    if pd.isna(a12):
        return "n/a"
    d = abs(a12 - 0.5)
    if d < 0.06: return "negligible"
    if d < 0.14: return "small"
    if d < 0.21: return "medium"
    return "large"


def benjamini_hochberg(pvals):
    p = np.asarray(pvals, dtype=float)
    mask = ~np.isnan(p)
    adj = np.full_like(p, np.nan, dtype=float)
    if mask.sum() == 0:
        return adj
    valid = p[mask]
    n = len(valid)
    order = np.argsort(valid)
    ranked = valid[order]
    q = ranked * n / np.arange(1, n + 1)
    q = np.minimum.accumulate(q[::-1])[::-1]
    adj_valid = np.empty(n)
    adj_valid[order] = np.minimum(q, 1.0)
    adj[mask] = adj_valid
    return adj


def paired_wilcoxon(x, y):
    x = np.asarray(x, dtype=float); y = np.asarray(y, dtype=float)
    mask = ~(np.isnan(x) | np.isnan(y))
    x, y = x[mask], y[mask]
    if len(x) < 6:
        return np.nan, np.nan, f"n={len(x)} < 6"
    diffs = x - y
    if np.allclose(diffs, 0):
        return 0.0, 1.0, "all pairs identical"
    try:
        res = wilcoxon(x, y, zero_method="wilcox", alternative="two-sided")
        return float(res.statistic), float(res.pvalue), ""
    except ValueError as e:
        return np.nan, np.nan, f"error: {e}"


# ─── Path discovery ────────────────────────────────────────────────────
def find_project_root():
    candidates = [Path.cwd()]
    candidates.extend(Path.cwd().parents)
    candidates.append(Path(__file__).resolve().parent)
    candidates.extend(Path(__file__).resolve().parents)
    seen = set()
    for c in candidates:
        c = c.resolve()
        if c in seen:
            continue
        seen.add(c)
        if (c / "files" / "Cases").exists():
            return c
    return None


# ─── Per-cell shard medians ────────────────────────────────────────────
def shard_medians(per_shard_path, sheet_name):
    """For one (SPL, approach x level) sheet, return a dict
    {shard_id: {column: median_value}} computing per-shard median across
    runs. Skips shards with HandledProducts == 0 (incomplete)."""
    try:
        df = pd.read_excel(per_shard_path, sheet_name=sheet_name)
    except Exception:
        return None
    if df.empty or "Shard" not in df.columns:
        return None

    numeric = df.select_dtypes(include=[np.number])
    g = df.groupby("Shard")[list(numeric.columns)].median()
    if "HandledProducts" in g.columns:
        g = g[g["HandledProducts"] > 0]
    if g.empty:
        return None

    out = {}
    for shard, row in g.iterrows():
        rec = {c: float(row[c]) for c in g.columns}
        # Derive per-shard throughput in products/second from the
        # canonical schema. HandledProducts per shard is fixed for
        # uncapped runs but can vary across approaches when the
        # 6-hour timeout binds (Hockerty Shirts), which is precisely
        # why this metric is more informative than T_pipeline alone.
        hp = rec.get("HandledProducts")
        tp_ms = rec.get("T_pipeline(ms)")
        if hp is not None and tp_ms is not None and tp_ms > 0:
            rec["__throughput__"] = float(hp) / (float(tp_ms) / 1000.0)
        else:
            rec["__throughput__"] = float("nan")
        out[int(shard)] = rec
    return out


# ─── Per-cell pairwise comparison ──────────────────────────────────────
def compare_one_cell(per_shard_path, spl_short, pairing):
    """Return list of one row per metric for this (SPL, pairing)."""
    appA, lvlA, appB, lvlB, label = pairing
    sheetA = f"{appA}_{lvlA}"
    sheetB = f"{appB}_{lvlB}"

    dataA = shard_medians(per_shard_path, sheetA)
    dataB = shard_medians(per_shard_path, sheetB)
    if dataA is None or dataB is None:
        return []

    common_shards = sorted(set(dataA) & set(dataB))
    if not common_shards:
        return []

    out_rows = []
    for canonical_col, metric_label, lower_is_better in METRICS:
        if canonical_col == "__coverage__":
            colA, colB, metric_label_resolved = resolve_coverage_col(appA, lvlA, appB, lvlB)
            if colA is None:
                continue
            metric_label = metric_label_resolved
        elif canonical_col == "__throughput__":
            # Derived per-shard throughput stored under '__throughput__'
            # in the shard-medians dict.
            colA = colB = "__throughput__"
        else:
            colA = colB = canonical_col

        x = np.array([dataA[s].get(colA, np.nan) for s in common_shards])
        y = np.array([dataB[s].get(colB, np.nan) for s in common_shards])
        valid = ~(np.isnan(x) | np.isnan(y))
        x_v, y_v = x[valid], y[valid]
        n = len(x_v)

        W, p, note = paired_wilcoxon(x_v, y_v)
        a12 = vargha_delaney_a12(x_v, y_v)
        mag = a12_magnitude(a12)

        med_a = float(np.median(x_v)) if n else np.nan
        med_b = float(np.median(y_v)) if n else np.nan
        med_delta = float(np.median(x_v - y_v)) if n else np.nan

        # Winner direction depends on metric semantics
        if pd.isna(a12):
            winner = "n/a"
        elif lower_is_better:
            winner = APPROACH_LABEL[appA] if a12 < 0.5 else (
                     APPROACH_LABEL[appB] if a12 > 0.5 else "tie")
        else:
            winner = APPROACH_LABEL[appA] if a12 > 0.5 else (
                     APPROACH_LABEL[appB] if a12 < 0.5 else "tie")

        out_rows.append({
            "SPL":                       spl_short,
            "Pairing":                   label,
            "ApproachA":                 APPROACH_LABEL[appA],
            "LevelA":                    lvlA,
            "ApproachB":                 APPROACH_LABEL[appB],
            "LevelB":                    lvlB,
            "Metric":                    metric_label,
            "N_shards":                  n,
            "median_A":                  round(med_a, 4) if not pd.isna(med_a) else np.nan,
            "median_B":                  round(med_b, 4) if not pd.isna(med_b) else np.nan,
            "median_delta_A_minus_B":    round(med_delta, 4) if not pd.isna(med_delta) else np.nan,
            "A12_A_vs_B":                round(a12, 4) if not pd.isna(a12) else np.nan,
            "magnitude":                 mag,
            "winner":                    winner,
            "W":                         W,
            "p":                         p,
            "note":                      note,
        })
    return out_rows


# ─── openpyxl: header style + auto-fit ─────────────────────────────────
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FONT = Font(bold=True, name='Calibri', size=11)
HEADER_FILL = PatternFill('solid', fgColor='E0E0E0')


def auto_fit_workbook(path, max_width=42):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        for col_idx in range(1, ws.max_column + 1):
            c = ws.cell(1, col_idx)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.border = THIN
            c.alignment = Alignment(horizontal='left', vertical='center')
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(1, col_idx).value
            best = len(str(header_val)) if header_val else 8
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, col_idx).value
                if v is not None:
                    best = max(best, len(str(v)))
            width = min(best * 1.1 + 2, max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)


# ─── Main ──────────────────────────────────────────────────────────────
def main():
    print("=" * 72)
    print("rq2_04 -- Paired Wilcoxon on 80 shards (per SPL x pairing)")
    print("=" * 72)

    project_root = find_project_root()
    if not project_root:
        print("ERROR: could not find project root containing files/Cases/.")
        sys.exit(1)
    cases_dir = project_root / "files" / "Cases"
    out_dir = project_root / "files" / "scripts" / "statistical_test_scripts" / "rq2_result"
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Project root: {project_root}")
    print(f"Output dir  : {out_dir}\n")

    all_rows = []
    for spl_folder, spl_short in SPL_MAPPING.items():
        per_shard_path = cases_dir / spl_folder / f"RQ2_perShard_{spl_folder}.xlsx"
        if not per_shard_path.exists():
            print(f"  [SKIP] {spl_short}: per-shard Excel missing")
            continue
        rows_for_spl = []
        for pairing in LEVEL_PAIRINGS:
            rows_for_spl.extend(compare_one_cell(per_shard_path, spl_short, pairing))
        if rows_for_spl:
            n_large = sum(1 for r in rows_for_spl if r["magnitude"] == "large")
            print(f"  [OK]  {spl_short:5s} : {len(rows_for_spl)} cells, "
                  f"{n_large} with large effect")
        all_rows.extend(rows_for_spl)

    if not all_rows:
        print("\nERROR: no comparisons produced.")
        sys.exit(1)

    df = pd.DataFrame(all_rows)

    # BH correction within each metric (the comparison family)
    df["p_BH"] = np.nan
    for metric in df["Metric"].unique():
        idx = df["Metric"] == metric
        mask_p = df.loc[idx, "p"].notna()
        if mask_p.any():
            df.loc[idx & mask_p, "p_BH"] = benjamini_hochberg(
                df.loc[idx & mask_p, "p"].values)
    df["significant_BH"] = df["p_BH"] < 0.05

    # SPL row order
    df["_spl_order"] = df["SPL"].map({s: i for i, s in enumerate(SPL_ORDER)})
    df = (df.sort_values(["_spl_order", "Pairing", "Metric"])
            .drop(columns=["_spl_order"])
            .reset_index(drop=True))

    col_order = [
        "SPL", "Pairing", "ApproachA", "LevelA", "ApproachB", "LevelB",
        "Metric", "N_shards", "median_A", "median_B", "median_delta_A_minus_B",
        "A12_A_vs_B", "magnitude", "winner",
        "W", "p", "p_BH", "significant_BH", "note",
    ]
    df = df[col_order]

    # ── Excel output ──
    out_xlsx = out_dir / "rq2_pairwise_wilcoxon.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="all_comparisons", index=False)
        for metric in df["Metric"].unique():
            sub = df[df["Metric"] == metric]
            if not sub.empty:
                # Excel sheet name max length 31, no special chars
                safe = metric.replace("(", "_").replace(")", "")[:31]
                sub.to_excel(writer, sheet_name=safe, index=False)
    auto_fit_workbook(out_xlsx)
    print(f"\nSaved: {out_xlsx.relative_to(project_root)}")

    # Console summary by metric
    print("\n--- Summary per metric ---")
    for metric in df["Metric"].unique():
        sub = df[df["Metric"] == metric]
        n = len(sub)
        large = int((sub["magnitude"] == "large").sum())
        sig = int(sub["significant_BH"].sum()) if sub["significant_BH"].notna().any() else 0
        print(f"  [{metric:24s}]  cells={n:3d}  large-effect={large:3d}  BH-sig={sig:3d}")

    print("\nrq2_04 DONE.")


if __name__ == "__main__":
    main()