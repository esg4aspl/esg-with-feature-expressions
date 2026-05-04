#!/usr/bin/env python3
"""
rq2_07_run_stability.py  (RQ2 Step 7 -- reviewer armor)

Quantifies run-to-run dispersion of RQ2 measurements. Reviewers will
ask: "You ran 11 reps on small/medium SPLs but only 3 on the industrial
ones (Tesla, syngo.via, HockertyShirts) -- how confident are the
medians?" This script answers that with per-shard coefficient of
variation (CV%) and relative IQR computed across runs.

Methodology
-----------
For each (SPL x approach x level x shard):
    runs    = repetitions of that exact partition (typically 11 or 3)
    CV%     = 100 * std(runs) / mean(runs)
    IQR%    = 100 * IQR(runs) / median(runs)

Both express dispersion on the same scale across SPLs of vastly
different absolute magnitudes (T_pipeline can range from milliseconds
on SVM to hours on HockertyShirts). The aggregate per (SPL x approach
x level x metric) is the median of these per-shard dispersion values
across the 80 shards.

Stability bands (qualitative, used in the 'Stability' column):
    median_CV% < 5%         "very stable"  -- 3 reps amply sufficient
    median_CV% in [5,15)%   "stable"       -- 3 reps acceptable
    median_CV% >= 15%       "notable variance"

Metrics covered (canonical schema)
----------------------------------
    T_pipeline(ms)             : end-to-end pipeline cost
    TestGenTime(ms)            : approach-specific generation cost
    TestGenPeakMemory(MB)      : peak heap during test generation
    EdgeCoverage(%) /          : effectiveness (uses whichever exists
    EventCoverage(%)              for the sheet's approach)

Output
------
files/scripts/statistical_test_scripts/rq2_result/rq2_run_stability.xlsx
    - 'summary'           : one row per (SPL x approach x level x metric)
                            with median/max CV%, median IQR%, Stability tag
    - 'industrial_focus'  : same, restricted to Tesla/syngo.via/HockertyShirts
                            (the 3-rep cells reviewers care about)
    - one detail sheet per metric : per-shard CV%/IQR% for drill-down
"""
from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ─── Configuration ─────────────────────────────────────────────────────
SPL_MAPPING = {
    "SodaVendingMachine":      "SVM",
    "eMail":                   "eM",
    "Elevator":                "El",
    "BankAccountv2":           "BAv2",
    "StudentAttendanceSystem": "SAS",
    "syngovia":                "Svia",
    "Tesla":                   "Te",
    "HockertyShirts":          "HS",
}
SPL_ORDER = ["SVM", "eM", "El", "BAv2", "SAS", "Svia", "Te", "HS"]

LARGE_SPLS = {"Tesla", "syngovia", "HockertyShirts"}

APPROACH_LABEL = {
    "ESG-Fx":     "Model Once, Generate Any",
    "EFG":        "Structural Baseline",
    "RandomWalk": "Stochastic Baseline",
}

# Metrics to evaluate. Coverage column is selected per-sheet (see below).
METRIC_COLS = [
    "T_pipeline(ms)",
    "TestGenTime(ms)",
    "TestGenPeakMemory(MB)",
    "__coverage__",
]


def coverage_col_for_sheet(sheet_name, df_cols):
    """Pick the coverage column appropriate for this sheet:
       - ESG-Fx_L1                  : EventCoverage(%)
       - ESG-Fx_L>=2                : EdgeCoverage(%)
       - EFG / RandomWalk           : EdgeCoverage(%) (the manuscript
                                        primary metric; both columns
                                        exist for these sheets)
    Returns the actual column name to use, or None.
    """
    if sheet_name == "ESG-Fx_L1":
        return "EventCoverage(%)" if "EventCoverage(%)" in df_cols else None
    return "EdgeCoverage(%)" if "EdgeCoverage(%)" in df_cols else None


# ─── Path discovery ────────────────────────────────────────────────────
def find_project_root():
    candidates = [Path.cwd()]
    candidates.extend(Path.cwd().parents)
    candidates.append(Path(__file__).resolve().parent)
    candidates.extend(Path(__file__).resolve().parents)
    seen = set()
    for c in candidates:
        c = c.resolve()
        if c in seen:
            continue
        seen.add(c)
        if (c / "files" / "Cases").exists():
            return c
    return None


def approach_from_sheet(sheet):
    if sheet.startswith("ESG-Fx"):
        return "ESG-Fx"
    if sheet.startswith("EFG"):
        return "EFG"
    if sheet.startswith("Random"):
        return "RandomWalk"
    return None


def level_from_sheet(sheet):
    return sheet.rsplit("_", 1)[-1] if "_" in sheet else ""


# ─── Per-shard dispersion ──────────────────────────────────────────────
def shard_dispersion_for_metric(df, metric_col):
    """For one sheet's perShard dataframe, compute per-shard CV% and
    IQR% across the available runs of the given metric. Skips shards
    with HandledProducts == 0 (incomplete) and shards with fewer than
    2 runs.

    Returns a list of dicts (one per shard).
    """
    if metric_col not in df.columns or "Shard" not in df.columns:
        return []

    out = []
    for shard, g in df.groupby("Shard"):
        if "HandledProducts" in g.columns and g["HandledProducts"].median() == 0:
            continue
        vals = g[metric_col].dropna().values
        n = len(vals)
        if n < 2:
            continue

        mean = float(np.mean(vals))
        std = float(np.std(vals, ddof=1))
        median = float(np.median(vals))
        q1, q3 = np.percentile(vals, [25, 75])
        iqr = float(q3 - q1)

        # CV is meaningless when mean is zero (e.g. coverage at 100%
        # with zero variance). Treat as 0 dispersion.
        cv_pct = (std / mean * 100.0) if mean != 0 else 0.0
        # Same for IQR%.
        iqr_pct = (iqr / median * 100.0) if median != 0 else 0.0

        out.append({
            "Shard":             int(shard),
            "N_runs":            n,
            "mean":              round(mean, 4),
            "median":            round(median, 4),
            "std":               round(std, 4),
            "IQR":               round(iqr, 4),
            "CV_pct":            round(cv_pct, 3),
            "IQR_pct_of_median": round(iqr_pct, 3),
        })
    return out


def stability_band(median_cv_pct):
    if median_cv_pct < 5:
        return "very stable"
    if median_cv_pct < 15:
        return "stable"
    return "notable variance"


# ─── openpyxl: header style + auto-fit ─────────────────────────────────
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FONT = Font(bold=True, name='Calibri', size=11)
HEADER_FILL = PatternFill('solid', fgColor='E0E0E0')


def auto_fit_workbook(path, max_width=42):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        for col_idx in range(1, ws.max_column + 1):
            c = ws.cell(1, col_idx)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.border = THIN
            c.alignment = Alignment(horizontal='left', vertical='center')
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(1, col_idx).value
            best = len(str(header_val)) if header_val else 8
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, col_idx).value
                if v is not None:
                    best = max(best, len(str(v)))
            width = min(best * 1.1 + 2, max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)


# ─── Main ──────────────────────────────────────────────────────────────
def main():
    print("=" * 72)
    print("rq2_07 -- Run-to-run stability  (CV% / IQR% per shard)")
    print("=" * 72)

    project_root = find_project_root()
    if not project_root:
        print("ERROR: could not find project root containing files/Cases/.")
        sys.exit(1)
    cases_dir = project_root / "files" / "Cases"
    out_dir = project_root / "files" / "scripts" / "statistical_test_scripts" / "rq2_result"
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Project root: {project_root}")
    print(f"Output dir  : {out_dir}\n")

    detail_rows = []
    summary_rows = []

    for spl_folder, spl_short in SPL_MAPPING.items():
        per_shard_path = cases_dir / spl_folder / f"RQ2_perShard_{spl_folder}.xlsx"
        if not per_shard_path.exists():
            print(f"  [SKIP] {spl_short}: per-shard Excel missing")
            continue
        is_large = spl_folder in LARGE_SPLS
        expected_runs = 3 if is_large else 11

        xls = pd.ExcelFile(per_shard_path)
        print(f"\n[{spl_short}] {spl_folder}  (expected runs: {expected_runs})")

        for sheet in xls.sheet_names:
            approach = approach_from_sheet(sheet)
            level = level_from_sheet(sheet)
            if approach is None:
                continue
            df = pd.read_excel(xls, sheet_name=sheet)

            for metric_col in METRIC_COLS:
                if metric_col == "__coverage__":
                    actual_col = coverage_col_for_sheet(sheet, df.columns)
                    if actual_col is None:
                        continue
                    metric_label = "Coverage(%)"
                else:
                    actual_col = metric_col
                    metric_label = metric_col

                shard_rows = shard_dispersion_for_metric(df, actual_col)
                if not shard_rows:
                    continue

                for r in shard_rows:
                    detail_rows.append({
                        "SPL":          spl_short,
                        "Scale":        "Industrial" if is_large else "Small/Medium",
                        "Approach":     APPROACH_LABEL[approach],
                        "Level":        level,
                        "Metric":       metric_label,
                        "MetricColumn": actual_col,
                        **r,
                    })

                cvs = [r["CV_pct"] for r in shard_rows]
                iqrs = [r["IQR_pct_of_median"] for r in shard_rows]
                med_cv = float(np.median(cvs))
                summary_rows.append({
                    "SPL":              spl_short,
                    "Scale":            "Industrial" if is_large else "Small/Medium",
                    "Approach":         APPROACH_LABEL[approach],
                    "Level":            level,
                    "Metric":           metric_label,
                    "MetricColumn":     actual_col,
                    "N_shards":         len(shard_rows),
                    "Expected_runs":    expected_runs,
                    "Median_CV_pct":    round(med_cv, 3),
                    "Max_CV_pct":       round(float(np.max(cvs)), 3),
                    "Median_IQR_pct":   round(float(np.median(iqrs)), 3),
                    "Stability":        stability_band(med_cv),
                })

    if not summary_rows:
        print("\nERROR: no stability data produced.")
        sys.exit(1)

    summary_df = pd.DataFrame(summary_rows)
    detail_df = pd.DataFrame(detail_rows)

    # SPL order
    summary_df["_spl_order"] = summary_df["SPL"].map(
        {s: i for i, s in enumerate(SPL_ORDER)})
    summary_df = (summary_df.sort_values(["_spl_order", "Approach", "Level", "Metric"])
                            .drop(columns=["_spl_order"])
                            .reset_index(drop=True))

    summary_cols = ["SPL", "Scale", "Approach", "Level", "Metric", "MetricColumn",
                    "N_shards", "Expected_runs",
                    "Median_CV_pct", "Max_CV_pct", "Median_IQR_pct", "Stability"]
    summary_df = summary_df[summary_cols]

    industrial_df = summary_df[summary_df["Scale"] == "Industrial"].reset_index(drop=True)

    # ── Excel output ──
    out_xlsx = out_dir / "rq2_run_stability.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        if not industrial_df.empty:
            industrial_df.to_excel(writer, sheet_name="industrial_focus", index=False)
        # Detail per metric
        for metric in detail_df["Metric"].unique():
            sub = detail_df[detail_df["Metric"] == metric]
            safe = metric.replace("(", "_").replace(")", "").replace("%", "pct")[:31]
            sub.to_excel(writer, sheet_name=f"detail_{safe}"[:31], index=False)
    auto_fit_workbook(out_xlsx)
    print(f"\nSaved: {out_xlsx.relative_to(project_root)}")

    # Console: industrial SPL focus (the reviewer question)
    print("\n--- Industrial SPL stability (reviewer focus) ---")
    if industrial_df.empty:
        print("  (no industrial SPL data)")
    else:
        for _, row in industrial_df.iterrows():
            print(f"  {row['SPL']:5s} {row['Approach']:30s} {row['Level']:3s} "
                  f"{row['Metric']:20s} median_CV={row['Median_CV_pct']:6.2f}% "
                  f"[{row['Stability']}]")

    flagged = summary_df[summary_df["Stability"] == "notable variance"]
    if not flagged.empty:
        print("\n--- Cells flagged as 'notable variance' (median CV >= 15%) ---")
        for _, row in flagged.iterrows():
            print(f"  {row['SPL']:5s} {row['Approach']:30s} {row['Level']:3s} "
                  f"{row['Metric']:20s} median_CV={row['Median_CV_pct']:6.2f}%")

    print("\nrq2_07 DONE.")


if __name__ == "__main__":
    main()