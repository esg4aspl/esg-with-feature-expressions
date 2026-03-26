#!/usr/bin/env python3
"""
rq5_main_automatic.py — Automatic RQ5 Test Suite Redundancy Analysis
======================================================================
Automatically discovers all SPLs and runs RQ5 analysis. NO ARGUMENTS NEEDED.

Usage:
    python rq5_main_automatic.py

Or with options:
    python rq5_main_automatic.py --approaches ALL
"""

import os
import sys
import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from rq5_readers import read_product_configurations, discover_all_suites
from rq5_analysis import (
    analyze_approach,
    compute_unique_sequence_ratios,
    compute_cumulative_edge_coverage,
)
from rq5_plots import (
    plot_scatter_per_approach,
    plot_cumulative_edge_coverage,
    plot_unique_sequence_ratios,
)

# =============================================================================
# SPL NAME MAPPING
# =============================================================================

SPL_NAME_MAPPING = {
    #"SodaVendingMachine": "SVM",
    #"eMail": "eM",
    #"Elevator": "El",
    #"BankAccountv2": "BAv2",
    #"StudentAttendanceSystem": "SAS",
    "Tesla": "Te",
    "syngovia": "Svia",
    "HockertyShirts": "HS"
}


# =============================================================================
# PROJECT ROOT FINDER
# =============================================================================

def find_project_root():
    """Find project root by looking for 'files/Cases' directory."""
    current = Path.cwd()
    
    # Try current directory first
    if (current / "files" / "Cases").exists():
        return current
    
    # Try one level up
    if (current.parent / "files" / "Cases").exists():
        return current.parent
    
    # Try two levels up
    if (current.parent.parent / "files" / "Cases").exists():
        return current.parent.parent
    
    # If we're in files/scripts, go up two levels
    if current.name == "scripts" and current.parent.name == "files":
        return current.parent.parent
    
    return None


# =============================================================================
# SPL DISCOVERY
# =============================================================================

def discover_spls_with_configurations(project_root):
    """
    Automatically discover all SPLs that have product configuration files.
    
    Returns: List of (spl_name, base_path, config_filename) tuples
    """
    cases_dir = project_root / "files" / "Cases"
    
    if not cases_dir.exists():
        print(f"❌ ERROR: Cases directory not found: {cases_dir}")
        return []
    
    spls = []
    
    for spl_dir in sorted(cases_dir.iterdir()):
        if not spl_dir.is_dir():
            continue
        
        # Look for configuration file - try both folder name and file prefix
        spl_file_prefix = SPL_NAME_MAPPING.get(spl_dir.name, spl_dir.name)
        
        config_patterns = [
            f"{spl_dir.name}_ProductConfigurations.txt",
            f"{spl_file_prefix}_ProductConfigurations.txt",  # Try file prefix
            "ProductConfigurations.txt",
            f"{spl_dir.name}_Configurations.txt",
            f"{spl_file_prefix}_Configurations.txt",  # Try file prefix
        ]
        
        config_file = None
        for pattern in config_patterns:
            candidate = spl_dir / pattern
            if candidate.exists():
                config_file = pattern
                break
        
        if config_file:
            # Check if test sequences exist
            ts_dir = spl_dir / "testsequences"
            if ts_dir.exists() and any(ts_dir.iterdir()):
                spls.append((spl_dir.name, str(spl_dir), config_file))
    
    return spls


# =============================================================================
# EXCEL STYLING
# =============================================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
CELL_FONT = Font(name='Arial', size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def _style_header(ws, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def _style_data(ws, n_rows, n_cols):
    for row in range(2, n_rows + 2):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            if isinstance(cell.value, float):
                cell.number_format = '0.0000'


def _auto_width(ws, df):
    for idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), 10)
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = max_len + 3
    ws.freeze_panes = 'A2'


# =============================================================================
# EXCEL WRITERS
# =============================================================================

def write_pairwise_excel(df, correlations, output_path):
    """Write per-approach pairwise results to Excel."""
    wb = Workbook()

    # Sheet 1: Pairwise data
    ws1 = wb.active
    ws1.title = 'Pairwise'
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)
    _style_header(ws1, len(df.columns))
    _style_data(ws1, len(df), len(df.columns))
    _auto_width(ws1, df)

    # Sheet 2: Correlations
    ws2 = wb.create_sheet('Correlations')
    ws2.append(['Similarity_Type', 'Spearman_rho', 'Spearman_p', 'Pearson_r', 'Pearson_p', 'N_pairs'])
    for metric, vals in correlations.items():
        ws2.append([metric, vals['spearman_rho'], vals['spearman_p'],
                     vals['pearson_r'], vals['pearson_p'], vals['n']])
    _style_header(ws2, 6)

    # Sheet 3: Descriptive statistics
    ws3 = wb.create_sheet('Descriptive')
    desc = df[['Config_Jaccard', 'Sequence_Jaccard', 'Edge_Jaccard']].describe().reset_index()
    for r in dataframe_to_rows(desc, index=False, header=True):
        ws3.append(r)
    _style_header(ws3, len(desc.columns))

    wb.save(output_path)


def write_summary_excel(all_results, unique_ratios_all, cumulative_all, output_path, spl_name):
    """Write cross-approach summary to Excel."""
    wb = Workbook()

    # Sheet 1: Correlation comparison
    ws1 = wb.active
    ws1.title = 'Correlations'
    headers = ['Approach', 'Similarity_Type', 'Spearman_rho', 'Spearman_p', 'Pearson_r', 'Pearson_p', 'N_pairs']
    ws1.append(headers)
    for approach, data in all_results.items():
        for metric, vals in data['correlations'].items():
            ws1.append([approach, metric, vals['spearman_rho'], vals['spearman_p'],
                         vals['pearson_r'], vals['pearson_p'], vals['n']])
    _style_header(ws1, len(headers))

    # Sheet 2: Descriptive stats per approach
    ws2 = wb.create_sheet('DescriptiveStats')
    ws2.append(['Approach', 'Metric', 'Mean', 'Median', 'Std', 'Min', 'Max', 'N_products', 'N_pairs'])
    for approach, data in all_results.items():
        df = data['dataframe']
        for col in ['Config_Jaccard', 'Sequence_Jaccard', 'Edge_Jaccard']:
            ws2.append([approach, col, df[col].mean(), df[col].median(), df[col].std(),
                         df[col].min(), df[col].max(), data['n_products'], data['n_pairs']])
    _style_header(ws2, 9)

    # Sheet 3: Unique Sequence Ratios
    ws3 = wb.create_sheet('UniqueSequenceRatio')
    ws3.append(['Approach', 'Product', 'Unique_Sequences', 'Total_Sequences', 'Unique_Ratio'])
    for approach, ratios in unique_ratios_all.items():
        for product, (unique, total, ratio) in sorted(ratios.items()):
            ws3.append([approach, product, unique, total, ratio])
    _style_header(ws3, 5)

    # Sheet 4: Cumulative Edge Coverage
    ws4 = wb.create_sheet('CumulativeEdgeCoverage')
    ws4.append(['Approach', 'Product_Index', 'Cumulative_Unique_Edges'])
    for approach, curve in cumulative_all.items():
        for idx, edges in curve:
            ws4.append([approach, idx, edges])
    _style_header(ws4, 3)

    wb.save(output_path)


# =============================================================================
# PROCESS SINGLE SPL
# =============================================================================

DEFAULT_APPROACHES = ['RandomWalk', 'ESG-Fx_L1', 'ESG-Fx_L2', 'ESG-Fx_L3', 'ESG-Fx_L4']
ALL_APPROACHES = DEFAULT_APPROACHES + ['EFG_L2', 'EFG_L3', 'EFG_L4']

# Backward compatibility mapping
APPROACH_MAPPING = {
    'RandomWalk': 'ESG-Fx_L0',  # Map display name to internal name
}


def process_spl(spl_name, base_path, config_filename, approaches):
    """Process a single SPL for RQ5 analysis."""
    config_path = os.path.join(base_path, config_filename)

    print(f"\n{'='*80}")
    print(f"📊 RQ5 ANALYSIS: {spl_name}")
    print(f"{'='*80}")
    
    # Show file prefix if different from folder name
    spl_file_prefix = SPL_NAME_MAPPING.get(spl_name, spl_name)
    if spl_file_prefix != spl_name:
        print(f"  File prefix: {spl_file_prefix}")
    
    print(f"  Base path: {base_path}")
    print(f"  Config:    {config_filename}")

    # 1. Read configurations
    try:
        configs = read_product_configurations(config_path)
        print(f"  ✅ Loaded {len(configs)} product configurations")
    except Exception as e:
        print(f"  ❌ Failed to read configurations: {e}")
        return False

    if len(configs) < 2:
        print(f"  ⚠️  Only {len(configs)} products - need at least 2. Skipping.")
        return False

    # 2. Create output directory
    output_dir = os.path.join(base_path, 'testsequenceanalysis')
    os.makedirs(output_dir, exist_ok=True)

    # 3. Analyze each approach
    all_results = {}
    unique_ratios_all = {}
    cumulative_all = {}

    for approach in approaches:
        print(f"\n  {'─'*40}")
        print(f"  📂 {approach}")
        
        # Map display name to internal name if needed
        internal_approach = APPROACH_MAPPING.get(approach, approach)
        
        try:
            suites = discover_all_suites(base_path, internal_approach)
        except Exception as e:
            print(f"     ❌ Error loading suites: {e}")
            continue
        
        if len(suites) < 2:
            print(f"     ⚠️  Found {len(suites)} suites - need ≥ 2. Skipping.")
            continue

        print(f"     ✅ Loaded {len(suites)} test suites")
        
        try:
            result = analyze_approach(configs, suites, approach)
        except Exception as e:
            print(f"     ❌ Analysis failed: {e}")
            continue
        
        if result:
            all_results[approach] = result
            unique_ratios_all[approach] = result['unique_ratios']
            cumulative_all[approach] = result['cumulative_curve']

            # Write per-approach Excel
            safe = approach.replace('-', '').replace('_', '')
            excel_path = os.path.join(output_dir, f'{spl_name}_RQ5_Pairwise_{safe}.xlsx')
            
            try:
                write_pairwise_excel(result['dataframe'], result['correlations'], excel_path)
                print(f"     ✅ Excel: {os.path.basename(excel_path)}")
            except Exception as e:
                print(f"     ❌ Failed to write Excel: {e}")

    if not all_results:
        print(f"\n  ⚠️  No approaches had sufficient data. Skipping {spl_name}.")
        return False

    # 4. Cross-approach summary Excel
    print(f"\n  {'─'*40}")
    print(f"  📊 Cross-approach summary")
    
    summary_path = os.path.join(output_dir, f'{spl_name}_RQ5_Summary.xlsx')
    try:
        write_summary_excel(all_results, unique_ratios_all, cumulative_all, summary_path, spl_name)
        print(f"     ✅ Summary: {os.path.basename(summary_path)}")
    except Exception as e:
        print(f"     ❌ Failed to write summary: {e}")

    # 5. Generate plots
    print(f"\n  {'─'*40}")
    print(f"  📈 Generating plots")
    
    try:
        plot_scatter_per_approach(all_results, output_dir, spl_name)
        plot_cumulative_edge_coverage(cumulative_all, output_dir, spl_name)
        plot_unique_sequence_ratios(unique_ratios_all, output_dir, spl_name)
        print(f"     ✅ Plots saved")
    except Exception as e:
        print(f"     ❌ Plot generation failed: {e}")

    print(f"\n  ✅ Output directory: {output_dir}/")
    
    return True


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Automatically run RQ5 redundancy analysis on all SPLs'
    )
    parser.add_argument(
        '--approaches',
        type=str,
        default='DEFAULT',
        choices=['DEFAULT', 'ALL'],
        help='Approach set: DEFAULT (ESG-Fx L0-L4) or ALL (+ EFG L2-L4)'
    )
    
    args = parser.parse_args()
    
    approaches = ALL_APPROACHES if args.approaches == 'ALL' else DEFAULT_APPROACHES
    
    print("="*80)
    print("RQ5 TEST SUITE REDUNDANCY ANALYSIS - AUTOMATIC")
    print("="*80)
    
    # Find project root
    project_root = find_project_root()
    
    if project_root is None:
        print("❌ ERROR: Could not find project root!")
        print(f"Current directory: {Path.cwd()}")
        print("\nSearched for: files/Cases/ directory")
        sys.exit(1)
    
    print(f"Project root: {project_root}")
    print(f"Cases dir:    {project_root / 'files' / 'Cases'}")
    print(f"Approaches:   {', '.join(approaches)}")
    print()
    
    # Discover SPLs
    print("🔍 Discovering SPLs with product configurations...")
    spls = discover_spls_with_configurations(project_root)
    
    if not spls:
        print("❌ No SPLs with product configurations found!")
        sys.exit(1)
    
    print(f"✅ Found {len(spls)} SPLs with configurations:")
    for spl_name, _, config_file in spls:
        file_prefix = SPL_NAME_MAPPING.get(spl_name, spl_name)
        if file_prefix != spl_name:
            print(f"   - {spl_name:25s} ({config_file}, files: {file_prefix}_*)")
        else:
            print(f"   - {spl_name:25s} ({config_file})")
    
    # Process each SPL
    print("\n" + "="*80)
    print("PROCESSING SPLs")
    print("="*80)
    
    processed_count = 0
    skipped_count = 0
    
    for spl_name, base_path, config_filename in spls:
        try:
            success = process_spl(spl_name, base_path, config_filename, approaches)
            if success:
                processed_count += 1
            else:
                skipped_count += 1
        except Exception as e:
            print(f"\n❌ ERROR processing {spl_name}: {e}")
            import traceback
            traceback.print_exc()
            skipped_count += 1
    
    # Final summary
    print("\n" + "="*80)
    print("FINAL SUMMARY")
    print("="*80)
    print(f"✅ Successfully processed: {processed_count} SPLs")
    if skipped_count > 0:
        print(f"⚠️  Skipped (insufficient data): {skipped_count} SPLs")
    
    print("\nOutput files (PNG format):")
    print("  - <SPL>_RQ5_Pairwise_<approach>.xlsx")
    print("  - <SPL>_RQ5_Summary.xlsx")
    print("  - <SPL>_RQ5_Scatter_<approach>.png")
    print("  - <SPL>_RQ5_ScatterOverview.png")
    print("  - <SPL>_RQ5_CumulativeEdge.png")
    print("  - <SPL>_RQ5_UniqueRatio.png")
    print("\n" + "="*80)


if __name__ == '__main__':
    main()