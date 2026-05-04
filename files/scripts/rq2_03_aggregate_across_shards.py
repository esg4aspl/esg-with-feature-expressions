#!/usr/bin/env python3
"""
rq2_03_aggregate_across_shards.py  (Step 3 of 3)

Input : files/Cases/<SPL>/RQ2_summary_<SPL>.xlsx     (from Step 2)
Output: files/Cases/RQ2_SPLSummary_medians.xlsx           (article: L=1 dropped)
        files/Cases/RQ2_SPLSummary_ForPhDThesis.xlsx      (thesis : L=1 included)
        files/Cases/RQ2_ApproachSummary_medians.xlsx      (article)
        files/Cases/RQ2_ApproachSummary_ForPhDThesis.xlsx (thesis)

Aggregates the 80 per-shard medians of each (SPL, approach x level)
into TWO summary rows: a SERIAL one (1-core sequential CPU cost) and
a WALL-CLOCK one (80-shard parallel deployment latency). Both
summaries describe exactly the same units of work executed; they
differ only on the time axis and the throughput it implies.

Aggregation rules
-----------------
  Time(ms)        SERIAL = SUM   WALLCLOCK = MAX
  PeakMemory(MB)  SERIAL = MAX   WALLCLOCK = MAX
                  (peak is peak regardless of scheduling)
  Coverage(%)     SERIAL = WAVG by HandledProducts   (same in both)
  SafetyLimitAvg  SERIAL = WAVG by SafetyLimitHitCount (same in both)
  Counts          SERIAL = SUM   WALLCLOCK = SUM
                  (work done is work done; HandledProducts, vertices,
                   edges, test cases, test events, AbortedSequences,
                   SafetyLimitHitCount -- all SUM in both tables)

Throughput (derived)
--------------------
  Throughput_serial    = sum(HandledProducts) / (sum(T_pipeline)/1000)
  Throughput_wallclock = sum(HandledProducts) / (max(T_pipeline)/1000)

  Both numerators are the same SUM -- 80 shards in parallel still
  process the same total products, only the elapsed time changes.
  Throughput_wallclock / Throughput_serial gives the realised
  parallel speed-up (ideal = 80).

Shares (computed from sums in BOTH tables -- share of total CPU is
the only meaningful interpretation; "share of bottleneck shard" is
not a coherent quantity).
  SATShare(%)            = sum(SatTime)            / sum(T_pipeline) * 100
  TransformationShare(%) = sum(TransformationTime) / sum(TestGenTime) * 100
                            (ESG-Fx L>=2 only)

Output structure
----------------
Each master Excel has multiple sheets. Each sheet has TWO stacked
tables -- Serial (top) and Wall-clock (bottom) -- separated by a
section banner row.

  SPLSummary  : one sheet per SPL.
                Within each table, a separate [header,data,blank]
                block per (approach x level) -- column sets vary
                across approaches (EFG has Parsing, RW has
                SafetyLimit, ESG-Fx L>=2 has Transformation, etc.)
                so a single union table would be sparse.

  ApproachSummary : one sheet per (approach x level).
                    Within each table, a single header row + one
                    row per SPL -- clean uniform tables since all
                    SPLs share the same column set for a given
                    approach x level.

L=1 handling
------------
ESG-Fx_L1 is structurally distinct (single Euler cycle on the
original graph, no L-sequence transformation -- transfer note
section 4). L=1 is excluded from manuscript figures but included
in thesis figures. The 'medians' files drop ESG-Fx_L1; the
'ForPhDThesis' files include it.
"""

import pandas as pd
import numpy as np
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ─── Configuration ─────────────────────────────────────────────────────
SHEET_ORDER = [
    'ESG-Fx_L1',  # thesis only -- dropped in 'medians' output
    'ESG-Fx_L2', 'ESG-Fx_L3', 'ESG-Fx_L4',
    'EFG_L2', 'EFG_L3', 'EFG_L4',
    'RandomWalk_L0',
]

SPL_ORDER = [
    'SodaVendingMachine', 'eMail', 'Elevator', 'BankAccountv2',
    'StudentAttendanceSystem', 'syngovia', 'Tesla', 'HockertyShirts',
]

THESIS_ONLY_SHEETS = {'ESG-Fx_L1'}

# Per-hit averages -- weighted by SafetyLimitHitCount, not summed.
SAFETY_LIMIT_AVG_COLS = {
    'AvgTimeOnSafetyLimit(ms)',
    'AvgStepsOnSafetyLimit',
    'AvgCoverageAtSafetyLimit(%)',
}


# ─── openpyxl styling ──────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
HEADER_FONT = Font(bold=True, name='Calibri', size=11)
DATA_FONT = Font(name='Calibri', size=11)
SECTION_FONT = Font(bold=True, italic=True, name='Calibri', size=12)
HEADER_FILL = PatternFill('solid', fgColor='E0E0E0')
SECTION_FILL = PatternFill('solid', fgColor='FFE699')
HEADER_ALIGN = Alignment(horizontal='left', vertical='center')
DATA_ALIGN = Alignment(horizontal='left', vertical='center')


# ─── Path discovery ────────────────────────────────────────────────────
def find_project_root():
    """Pure relative discovery -- no hardcoded user paths."""
    candidates = [Path.cwd()]
    candidates.extend(Path.cwd().parents)
    candidates.append(Path(__file__).resolve().parent)
    candidates.extend(Path(__file__).resolve().parents)
    seen = set()
    for c in candidates:
        c = c.resolve()
        if c in seen:
            continue
        seen.add(c)
        if (c / "files" / "Cases").exists():
            return c
    return None


def auto_col_width(text):
    return max(len(str(text)) * 1.15 + 2, 8)


# ─── Aggregation primitives ────────────────────────────────────────────
def classify_column(col_name):
    """Return a coarse type tag for aggregation dispatch.

    Order matters:
      - safety-limit avgs first (they end with 'Time(ms)' or '(%)' but
        are conceptually per-hit averages, not durations or shares)
      - peak memory before counts
      - time before coverage(%) (no overlap, but explicit is safer)
    """
    if col_name in SAFETY_LIMIT_AVG_COLS:
        return 'wavg_safety'
    if 'PeakMemory' in col_name:
        return 'max'
    if col_name.endswith('Time(ms)') or col_name == 'T_pipeline(ms)':
        return 'time'
    if col_name.endswith('Coverage(%)'):
        return 'wavg_handled'
    return 'count'  # default: SUM (HandledProducts, FailedProducts,
                    # vertices, edges, test cases, test events,
                    # AbortedSequences, SafetyLimitHitCount, ...)


def safe_wavg(values, weights):
    """Weighted average that returns NaN cleanly if no positive weight
    or no valid data points exist (e.g. SafetyLimitHitCount=0 in every
    shard means the safety-limit averages are undefined, and we want
    a blank cell in the output rather than a misleading 0.0)."""
    weights = np.asarray(weights, dtype=float)
    values = np.asarray(values, dtype=float)
    mask = ~np.isnan(values) & ~np.isnan(weights) & (weights > 0)
    if not mask.any():
        return float('nan')
    w = weights[mask]
    v = values[mask]
    tw = w.sum()
    if tw <= 0:
        return float('nan')
    return float((v * w).sum() / tw)


def aggregate_one_sheet(df):
    """Collapse the 80 per-shard rows into TWO dicts (serial, wallclock).

    Returns:
        (serial, wallclock) dicts keyed by canonical column name.
        They share most keys (counts, memory, coverage, safety-limit
        avgs); they differ on time-related ones.
    """
    exclude = {'Shard', 'N_Runs', 'Approach', 'SPLName', 'CoverageType'}
    numeric_cols = [c for c in df.columns
                    if c not in exclude and pd.api.types.is_numeric_dtype(df[c])]

    serial, wallclock = {}, {}

    for c in numeric_cols:
        kind = classify_column(c)
        if kind == 'max':
            v = df[c].max()
            serial[c] = v
            wallclock[c] = v
        elif kind == 'wavg_safety':
            if 'SafetyLimitHitCount' in df.columns:
                w = df['SafetyLimitHitCount']
            else:
                w = df.get('HandledProducts', pd.Series([1] * len(df)))
            v = safe_wavg(df[c], w)
            serial[c] = v
            wallclock[c] = v
        elif kind == 'wavg_handled':
            w = df.get('HandledProducts', pd.Series([1] * len(df)))
            v = safe_wavg(df[c], w)
            serial[c] = v
            wallclock[c] = v
        elif kind == 'time':
            serial[c] = df[c].sum()
            wallclock[c] = df[c].max()
        else:  # count
            v = df[c].sum()
            serial[c] = v
            wallclock[c] = v

    # Carry constant labels (Approach, SPLName, CoverageType).
    for c in ('Approach', 'SPLName', 'CoverageType'):
        if c in df.columns and len(df) > 0:
            serial[c] = df[c].iloc[0]
            wallclock[c] = df[c].iloc[0]

    return serial, wallclock


def finalize_row(serial, wallclock):
    """Add derived columns (Throughput, SATShare, TransformationShare).

    Mutates serial and wallclock in place.
    """
    handled = serial.get('HandledProducts', 0) or 0
    t_serial = serial.get('T_pipeline(ms)', 0) or 0
    t_wall = wallclock.get('T_pipeline(ms)', 0) or 0

    serial['Throughput_serial(prod_per_sec)'] = (
        handled / (t_serial / 1000.0) if t_serial > 0 else float('nan'))
    wallclock['Throughput_wallclock(prod_per_sec)'] = (
        handled / (t_wall / 1000.0) if t_wall > 0 else float('nan'))

    # SATShare(%) -- share of total pipeline CPU spent in SAT solving;
    # computed from sums in BOTH tables (the same number is the
    # meaningful one in both contexts).
    sat = serial.get('SatTime(ms)', 0) or 0
    share = (sat / t_serial * 100.0) if t_serial > 0 else float('nan')
    serial['SATShare(%)'] = share
    wallclock['SATShare(%)'] = share

    # TransformationShare(%) -- only meaningful for ESG-Fx L>=2
    # (L=1 has no transformation; EFG/RW don't have it either).
    if 'TransformationTime(ms)' in serial:
        trans = serial.get('TransformationTime(ms)', 0) or 0
        testgen = serial.get('TestGenTime(ms)', 0) or 0
        ts = (trans / testgen * 100.0) if testgen > 0 else float('nan')
        serial['TransformationShare(%)'] = ts
        wallclock['TransformationShare(%)'] = ts


# ─── Output column ordering ────────────────────────────────────────────
def ordered_columns_for(row, leading_label_col, throughput_col):
    """Return the ordered column list to emit for one row.

    Drops any column not in 'row'. 'leading_label_col' is 'Approach'
    for SPLSummary (each row labelled by approach x level) or 'SPL'
    for ApproachSummary (each row labelled by SPL).
    """
    order = [
        leading_label_col, 'CoverageType',
        'HandledProducts', 'FailedProducts',
        'T_pipeline(ms)', throughput_col,
        'SatTime(ms)', 'SATShare(%)',
        'ProdGenTime(ms)',
        'EFGTransformationTime(ms)',
        'TransformationTime(ms)', 'TransformationShare(%)',
        'TestGenTime(ms)',
        'ParsingTime(ms)',
        'TestExecTime(ms)',
        'TestGenPeakMemory(MB)', 'TestExecPeakMemory(MB)',
        'NumberOfESGFxVertices', 'NumberOfESGFxEdges',
        'NumberOfESGFxTestCases', 'NumberOfESGFxTestEvents',
        'NumberOfEFGVertices', 'NumberOfEFGEdges',
        'NumberOfEFGTestCases', 'NumberOfEFGTestEvents',
        'EventCoverage(%)', 'EventCoverageAnalysisTime(ms)',
        'EdgeCoverage(%)', 'EdgeCoverageAnalysisTime(ms)',
        'AbortedSequences',
        'SafetyLimitHitCount',
        'AvgTimeOnSafetyLimit(ms)', 'AvgStepsOnSafetyLimit',
        'AvgCoverageAtSafetyLimit(%)',
        'Java_T_total_buggy(ms)',
    ]
    return [c for c in order if c in row]


# ─── openpyxl writers ──────────────────────────────────────────────────
def write_section_header(ws, row, text, ncols=20):
    cell = ws.cell(row, 1, text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = HEADER_ALIGN
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=ncols)


def write_header_row(ws, row, headers, max_widths):
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row, col_idx, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = THIN_BORDER
        max_widths[col_idx] = max(max_widths.get(col_idx, 0), auto_col_width(h))


def write_data_row(ws, row, headers, data, max_widths):
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row, col_idx)
        v = data.get(h)
        if v is None:
            c.value = None
        elif isinstance(v, (float, np.floating)):
            if np.isnan(v):
                c.value = None  # leave undefined cells blank
            else:
                c.value = round(float(v), 2)
        elif isinstance(v, (int, np.integer)):
            c.value = int(v)
        else:
            c.value = v
        c.font = DATA_FONT
        c.alignment = DATA_ALIGN
        c.border = THIN_BORDER
        s = "" if c.value is None else str(c.value)
        max_widths[col_idx] = max(max_widths.get(col_idx, 0), auto_col_width(s))


def apply_column_widths(ws, max_widths):
    for col_idx, w in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(w, 40)


# ─── SPLSummary writer (one sheet per SPL) ─────────────────────────────
def write_spl_sheet(ws, rows_per_approach, available_sheets):
    """SPLSummary: one sheet per SPL.

    Layout: SERIAL section (multiple [header,data,blank] approach
    blocks), then WALL-CLOCK section, same structure. Approach blocks
    are needed because column sets vary across approaches.
    """
    max_widths = {}
    cur = 1

    sections = (
        ('SERIAL summary  (1-core sequential CPU cost)',
         'Throughput_serial(prod_per_sec)', 0),
        ('WALL-CLOCK summary  (80-shard parallel deployment latency)',
         'Throughput_wallclock(prod_per_sec)', 1),
    )

    for label, throughput_col, side_key in sections:
        write_section_header(ws, cur, label, ncols=25)
        cur += 2  # banner + blank

        for sheet_name in available_sheets:
            row_pair = rows_per_approach[sheet_name]
            row_data = row_pair[side_key]
            headers = ordered_columns_for(row_data, 'Approach', throughput_col)
            write_header_row(ws, cur, headers, max_widths)
            write_data_row(ws, cur + 1, headers, row_data, max_widths)
            cur += 3  # header, data, blank

        cur += 1  # extra space before next section

    apply_column_widths(ws, max_widths)


# ─── ApproachSummary writer (one sheet per approach x level) ───────────
def write_approach_sheet(ws, rows_per_spl, ordered_spls):
    """ApproachSummary: one sheet per (approach x level).

    Layout: SERIAL section (single header + N SPL rows), then
    WALL-CLOCK section. Within each section, the column set is uniform
    across SPLs so the table is clean.
    """
    max_widths = {}
    cur = 1

    sections = (
        ('SERIAL summary  (1-core sequential CPU cost)',
         'Throughput_serial(prod_per_sec)', 0),
        ('WALL-CLOCK summary  (80-shard parallel deployment latency)',
         'Throughput_wallclock(prod_per_sec)', 1),
    )

    for label, throughput_col, side_key in sections:
        write_section_header(ws, cur, label, ncols=25)
        cur += 2

        # Build SPL rows in the requested order (skip SPLs with no data).
        rows_for_section = []
        for spl in ordered_spls:
            if spl not in rows_per_spl:
                continue
            row_data = dict(rows_per_spl[spl][side_key])
            row_data['SPL'] = spl
            rows_for_section.append(row_data)

        if not rows_for_section:
            cur += 1
            continue

        # Column union across SPLs (same column set in practice, but be
        # defensive in case some SPL is missing a column due to data).
        seen_cols = set()
        col_union = []
        for r in rows_for_section:
            for c in ordered_columns_for(r, 'SPL', throughput_col):
                if c not in seen_cols:
                    col_union.append(c)
                    seen_cols.add(c)

        write_header_row(ws, cur, col_union, max_widths)
        cur += 1
        for r in rows_for_section:
            write_data_row(ws, cur, col_union, r, max_widths)
            cur += 1
        cur += 2

    apply_column_widths(ws, max_widths)


# ─── Main ──────────────────────────────────────────────────────────────
def process_spl(filepath):
    """Return ({sheet_name: (serial, wallclock)}, ordered_sheet_list)."""
    xls = pd.ExcelFile(filepath)
    available = [s for s in SHEET_ORDER if s in xls.sheet_names]
    rows = {}
    for sheet_name in available:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        serial, wallclock = aggregate_one_sheet(df)
        finalize_row(serial, wallclock)
        # For SPLSummary display, the leading label is the full sheet
        # name (approach + level) so each block is unambiguous.
        serial['Approach'] = sheet_name
        wallclock['Approach'] = sheet_name
        # Make sure CoverageType is set even if the input lost it.
        if 'CoverageType' not in serial:
            serial['CoverageType'] = sheet_name.rsplit('_', 1)[-1]
            wallclock['CoverageType'] = sheet_name.rsplit('_', 1)[-1]
        rows[sheet_name] = (serial, wallclock)
    return rows, available


def write_spl_summary(out_path, all_spl_data, include_l1):
    wb = Workbook()
    first = True

    def emit(spl_name, rows, sheets):
        nonlocal first
        if not include_l1:
            sheets = [s for s in sheets if s not in THESIS_ONLY_SHEETS]
        if not sheets:
            return
        if first:
            ws = wb.active
            ws.title = spl_name
            first = False
        else:
            ws = wb.create_sheet(title=spl_name)
        write_spl_sheet(ws, rows, sheets)

    for spl_name in SPL_ORDER:
        if spl_name not in all_spl_data:
            continue
        rows, sheets = all_spl_data[spl_name]
        emit(spl_name, rows, sheets)

    # Defensive: any SPL not in SPL_ORDER still gets a sheet at the end.
    for spl_name, (rows, sheets) in all_spl_data.items():
        if spl_name in SPL_ORDER:
            continue
        emit(spl_name, rows, sheets)

    if first:
        wb.active.title = "empty"
    wb.save(out_path)


def write_approach_summary(out_path, all_spl_data, include_l1):
    """Pivot per-SPL data into per-approach sheets."""
    pivoted = {}
    for spl_name, (rows, sheets) in all_spl_data.items():
        for sheet_name in sheets:
            if not include_l1 and sheet_name in THESIS_ONLY_SHEETS:
                continue
            pivoted.setdefault(sheet_name, {})[spl_name] = rows[sheet_name]

    wb = Workbook()
    first = True
    for sheet_name in SHEET_ORDER:
        if sheet_name not in pivoted:
            continue
        rows_per_spl = pivoted[sheet_name]
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        write_approach_sheet(ws, rows_per_spl, SPL_ORDER)

    if first:
        wb.active.title = "empty"
    wb.save(out_path)


def main():
    print("="*80)
    print("RQ2 STEP 3 / 3 - AGGREGATE 80 SHARDS  (serial + wall-clock)")
    print("="*80)

    project_root = find_project_root()
    if not project_root:
        print("ERROR: Could not find project root containing files/Cases/.")
        sys.exit(1)

    cases_dir = project_root / "files" / "Cases"
    files = sorted(cases_dir.rglob("RQ2_summary_*.xlsx"))
    if not files:
        print("ERROR: No RQ2_summary_*.xlsx files found! Run Step 2 first.")
        sys.exit(1)

    print(f"Project root: {project_root}")
    print(f"Cases dir   : {cases_dir}")
    print(f"Found {len(files)} summary file(s)\n")

    all_spl_data = {}
    for f in files:
        spl_name = f.stem.replace("RQ2_summary_", "")
        print(f"Processing {spl_name}...")
        rows, sheets = process_spl(f)
        all_spl_data[spl_name] = (rows, sheets)
        print(f"  -> {len(sheets)} approach x level sheet(s)")

    out_spl_med = cases_dir / "RQ2_SPLSummary_medians.xlsx"
    out_spl_thes = cases_dir / "RQ2_SPLSummary_ForPhDThesis.xlsx"
    out_app_med = cases_dir / "RQ2_ApproachSummary_medians.xlsx"
    out_app_thes = cases_dir / "RQ2_ApproachSummary_ForPhDThesis.xlsx"

    write_spl_summary(out_spl_med, all_spl_data, include_l1=False)
    write_spl_summary(out_spl_thes, all_spl_data, include_l1=True)
    write_approach_summary(out_app_med, all_spl_data, include_l1=False)
    write_approach_summary(out_app_thes, all_spl_data, include_l1=True)

    print("\nMaster Excels written:")
    for p in (out_spl_med, out_spl_thes, out_app_med, out_app_thes):
        try:
            print(f"  {p.relative_to(project_root)}")
        except ValueError:
            print(f"  {p}")

    print("\nDONE.")


if __name__ == "__main__":
    main()