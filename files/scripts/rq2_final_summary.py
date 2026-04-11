#!/usr/bin/env python3
"""
rq2_final_summary.py

Aggregation rules for 80 shards -> 1 row per approach:
  * Time columns              -> SUM
  * Count columns             -> SUM (products, vertices, edges, test cases/events, etc.)
  * Peak Memory               -> MAX
  * Coverage (%)              -> Weighted average by Processed Products
  * Aborted Sequences, Safety Limit Hit Count, Avg Time/Steps on Safety Limit -> SUM
  * Avg Coverage at Safety Limit(%) -> Weighted average by Processed Products
"""

import pandas as pd
import numpy as np
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SHEET_ORDER = [
    'ESG-Fx_L1', 'ESG-Fx_L2', 'ESG-Fx_L3', 'ESG-Fx_L4',
    'EFG_L2', 'EFG_L3', 'EFG_L4',
    'RandomWalk_L0'
]

SPL_ORDER = [
    'SodaVendingMachine', 'eMail', 'Elevator', 'BankAccountv2',
    'StudentAttendanceSystem', 'syngovia', 'Tesla', 'HockertyShirts'
]

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FONT = Font(bold=True, name='Calibri', size=11)
DATA_FONT = Font(name='Calibri', size=11)
HEADER_ALIGN = Alignment(horizontal='left', vertical='center')
DATA_ALIGN = Alignment(horizontal='left', vertical='center')


def classify_column(col_name):
    if 'Peak Memory' in col_name:
        return 'max'
    if 'Coverage(%)' in col_name:
        return 'wavg'
    return 'sum'


def find_project_root():
    current = Path.cwd()
    for candidate in [current, current.parent, current.parent.parent]:
        if (candidate / "files" / "Cases").exists():
            return candidate
    hardcoded = Path("/Users/dilekozturk/git/esg-with-feature-expressions")
    if (hardcoded / "files" / "Cases").exists():
        return hardcoded
    return None


def auto_col_width(header_text):
    return max(len(str(header_text)) * 1.15 + 2, 8)


def summarize_one_sheet(df):
    exclude = {'Shard', 'N_Runs'}
    numeric_cols = [c for c in df.columns if c not in exclude and pd.api.types.is_numeric_dtype(df[c])]
    weight_col = 'Processed Products'
    weights = df[weight_col] if weight_col in df.columns else None

    result = {'Approach': ''}
    for c in numeric_cols:
        agg = classify_column(c)
        if agg == 'max':
            result[c] = df[c].max()
        elif agg == 'wavg' and weights is not None:
            tw = weights.sum()
            result[c] = (df[c] * weights).sum() / tw if tw > 0 else df[c].mean()
        else:
            result[c] = df[c].sum()
    return result


def process_spl(filepath):
    xls = pd.ExcelFile(filepath)
    available_sheets = [s for s in SHEET_ORDER if s in xls.sheet_names]
    rows_per_approach = {}
    for sheet_name in available_sheets:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        row = summarize_one_sheet(df)
        row['Approach'] = sheet_name
        rows_per_approach[sheet_name] = row
    return rows_per_approach, available_sheets


def write_approaches_to_sheet(ws, rows_per_approach, available_sheets):
    current_row = 1
    max_col_widths = {}
    for sheet_name in available_sheets:
        row_data = rows_per_approach[sheet_name]
        headers = list(row_data.keys())
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(current_row, col_idx, h)
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        for col_idx, h in enumerate(headers, 1):
            val = row_data[h]
            cell = ws.cell(current_row + 1, col_idx)
            if isinstance(val, (float, np.floating)):
                cell.value = round(val, 2)
            elif isinstance(val, (int, np.integer)):
                cell.value = int(val)
            else:
                cell.value = val
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
        for col_idx, h in enumerate(headers, 1):
            w = auto_col_width(h)
            max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), w)
        current_row += 3
    for col_idx, w in max_col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def main():
    project_root = find_project_root()
    if not project_root:
        print("Could not find project root.")
        sys.exit(1)

    cases_dir = project_root / "files" / "Cases"
    files = sorted(cases_dir.rglob("RQ2_summary_*.xlsx"))
    if not files:
        print("No RQ2_summary_*.xlsx files found!")
        sys.exit(1)

    print(f"Found {len(files)} summary file(s)\n")
    all_spl_data = {}

    for f in files:
        spl_name = f.stem.replace("RQ2_summary_", "")
        spl_dir = f.parent
        print(f"Processing {spl_name}...")
        rows_per_approach, available_sheets = process_spl(f)

        output_path = spl_dir / f"RQ2_final_summary_{spl_name}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = spl_name
        write_approaches_to_sheet(ws, rows_per_approach, available_sheets)
        wb.save(output_path)
        print(f"  -> {output_path}")
        all_spl_data[spl_name] = (rows_per_approach, available_sheets)

    global_path = cases_dir / "RQ2_SPL_Summary.xlsx"
    wb = Workbook()
    first = True
    for spl_name in SPL_ORDER:
        if spl_name not in all_spl_data:
            print(f"  [skip] {spl_name}")
            continue
        rows, sheets = all_spl_data[spl_name]
        if first:
            ws = wb.active
            ws.title = spl_name
            first = False
        else:
            ws = wb.create_sheet(title=spl_name)
        write_approaches_to_sheet(ws, rows, sheets)
    wb.save(global_path)
    print(f"\nGlobal summary: {global_path}")


if __name__ == "__main__":
    main()