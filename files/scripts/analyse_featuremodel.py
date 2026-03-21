#!/usr/bin/env python3
"""
analyse_featuremodel.py

Reads FeatureIDE-style feature model XML files from:
  /esg-with-feature-expressions/files/Cases/<CASE>/configs/model.xml

and produces:
  /esg-with-feature-expressions/files/Cases/FeatureModelAnalysis.xlsx

Columns:
  - SPL Name
  - Number of Features in Feature Model
  - Number of Concrete Features
  - Number of Constraints
  - Number of Possible Valid Configurations

Notes / assumptions:
- Supports common FeatureIDE XML structure tags: and, or, alt, feature
- Supports common constraint operators: var, not, conj, disj, imp, eq
- Counts configurations exactly using a SAT-style model counter implemented in pure Python
- "Number of Features in Feature Model" counts all named nodes under <struct>, including abstract ones and the root
- "Number of Concrete Features" counts named nodes whose attribute abstract!="true"
"""

from __future__ import annotations

import math
import os
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
import xml.etree.ElementTree as ET

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required. Install it with:\n"
        "  pip install openpyxl"
    ) from exc


# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------

BASE_CASES_DIR = Path("/esg-with-feature-expressions/files/Cases")
OUTPUT_XLSX = BASE_CASES_DIR / "FeatureModelAnalysis.xlsx"

CASES = [
    "SodaVendingMachine",
    "eMail",
    "Elevator",
    "BankAccountv2",
    "StudentAttendanceSystem",
    "Tesla",
    "syngovia",
    "HockertyShirts",
]


# ---------------------------------------------------------------------
# Feature model structure
# ---------------------------------------------------------------------

@dataclass
class FMNode:
    name: str
    kind: str                    # and / or / alt / feature
    abstract: bool = False
    mandatory: bool = False
    children: List["FMNode"] = field(default_factory=list)
    parent: Optional["FMNode"] = None

    def add_child(self, child: "FMNode") -> None:
        child.parent = self
        self.children.append(child)


# ---------------------------------------------------------------------
# CNF helper
# ---------------------------------------------------------------------

class CNFBuilder:
    def __init__(self) -> None:
        self.var_to_id: Dict[str, int] = {}
        self.id_to_var: Dict[int, str] = {}
        self.next_id = 1
        self.clauses: List[Tuple[int, ...]] = []
        self.aux_counter = 1

    def get_var(self, name: str) -> int:
        if name not in self.var_to_id:
            vid = self.next_id
            self.next_id += 1
            self.var_to_id[name] = vid
            self.id_to_var[vid] = name
        return self.var_to_id[name]

    def new_aux(self, prefix: str = "__aux") -> int:
        name = f"{prefix}_{self.aux_counter}"
        self.aux_counter += 1
        return self.get_var(name)

    def add_clause(self, *lits: int) -> None:
        lits = tuple(sorted(set(lits), key=lambda x: abs(x)))
        if 0 in lits:
            raise ValueError("Literal 0 is not allowed in CNF.")
        self.clauses.append(lits)

    def add_implication(self, a: int, b: int) -> None:
        # a -> b  ===  ¬a ∨ b
        self.add_clause(-a, b)

    def add_equivalence(self, a: int, b: int) -> None:
        # a <-> b  ===  (¬a ∨ b) ∧ (¬b ∨ a)
        self.add_clause(-a, b)
        self.add_clause(-b, a)


# ---------------------------------------------------------------------
# Parsing structure
# ---------------------------------------------------------------------

STRUCT_TAGS = {"and", "or", "alt", "feature"}


def clean_bool(value: Optional[str]) -> bool:
    return str(value).strip().lower() == "true"


def parse_struct_node(elem: ET.Element) -> FMNode:
    tag = elem.tag
    if tag not in STRUCT_TAGS:
        raise ValueError(f"Unsupported struct tag: {tag}")

    name = elem.attrib.get("name")
    if not name:
        raise ValueError(f"Unnamed feature/group node encountered: <{tag}>")

    node = FMNode(
        name=name,
        kind=tag,
        abstract=clean_bool(elem.attrib.get("abstract")),
        mandatory=clean_bool(elem.attrib.get("mandatory")),
    )

    for child in elem:
        if child.tag in STRUCT_TAGS:
            node.add_child(parse_struct_node(child))
        # ignore <graphics> etc.

    return node


def parse_feature_model(xml_path: Path) -> Tuple[FMNode, ET.Element]:
    tree = ET.parse(xml_path)
    root = tree.getroot()

    struct = root.find("struct")
    if struct is None:
        raise ValueError(f"No <struct> found in {xml_path}")

    struct_children = [c for c in struct if c.tag in STRUCT_TAGS]
    if len(struct_children) != 1:
        raise ValueError(
            f"Expected exactly one root feature under <struct> in {xml_path}, "
            f"found {len(struct_children)}"
        )

    fm_root = parse_struct_node(struct_children[0])
    return fm_root, root


# ---------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------

def collect_nodes(root: FMNode) -> List[FMNode]:
    result: List[FMNode] = []

    def dfs(node: FMNode) -> None:
        result.append(node)
        for child in node.children:
            dfs(child)

    dfs(root)
    return result


# ---------------------------------------------------------------------
# Structure semantics -> CNF
# ---------------------------------------------------------------------

def encode_structure(root: FMNode, cnf: CNFBuilder) -> None:
    # Root must be selected
    root_id = cnf.get_var(root.name)
    cnf.add_clause(root_id)

    def dfs(node: FMNode) -> None:
        p = cnf.get_var(node.name)

        # Parent-child relation from the parent's group type
        if node.kind == "and":
            for child in node.children:
                c = cnf.get_var(child.name)

                # child -> parent always
                cnf.add_implication(c, p)

                # mandatory child: parent -> child
                if child.mandatory:
                    cnf.add_implication(p, c)

                dfs(child)

        elif node.kind == "or":
            # each child -> parent
            child_ids = []
            for child in node.children:
                c = cnf.get_var(child.name)
                child_ids.append(c)
                cnf.add_implication(c, p)
                dfs(child)

            # parent -> at least one child
            if child_ids:
                cnf.add_clause(-p, *child_ids)

        elif node.kind == "alt":
            child_ids = []
            for child in node.children:
                c = cnf.get_var(child.name)
                child_ids.append(c)
                cnf.add_implication(c, p)
                dfs(child)

            # parent -> exactly one child
            if child_ids:
                cnf.add_clause(-p, *child_ids)
                for i in range(len(child_ids)):
                    for j in range(i + 1, len(child_ids)):
                        cnf.add_clause(-child_ids[i], -child_ids[j])

        elif node.kind == "feature":
            # leaf: semantics already handled by parent
            pass

        else:
            raise ValueError(f"Unsupported node kind: {node.kind}")

    dfs(root)


# ---------------------------------------------------------------------
# Constraint parsing
# ---------------------------------------------------------------------

def build_formula_var_set(elem: ET.Element, names: Set[str]) -> None:
    if elem.tag == "var":
        names.add(elem.text.strip())
        return
    for child in elem:
        build_formula_var_set(child, names)


def tseitin_encode_formula(elem: ET.Element, cnf: CNFBuilder) -> int:
    tag = elem.tag

    if tag == "var":
        if elem.text is None:
            raise ValueError("<var> without text")
        return cnf.get_var(elem.text.strip())

    if tag == "not":
        children = list(elem)
        if len(children) != 1:
            raise ValueError("<not> must have exactly one child")
        a = tseitin_encode_formula(children[0], cnf)
        x = cnf.new_aux("__not")
        # x <-> ¬a
        cnf.add_clause(-x, -a)
        cnf.add_clause(x, a)
        return x

    if tag == "conj":
        children = list(elem)
        if len(children) < 2:
            raise ValueError("<conj> must have at least two children")
        lits = [tseitin_encode_formula(ch, cnf) for ch in children]
        x = cnf.new_aux("__and")
        # x <-> (a1 ∧ ... ∧ an)
        for a in lits:
            cnf.add_clause(-x, a)
        cnf.add_clause(x, *[-a for a in lits])
        return x

    if tag == "disj":
        children = list(elem)
        if len(children) < 2:
            raise ValueError("<disj> must have at least two children")
        lits = [tseitin_encode_formula(ch, cnf) for ch in children]
        x = cnf.new_aux("__or")
        # x <-> (a1 ∨ ... ∨ an)
        cnf.add_clause(-x, *lits)
        for a in lits:
            cnf.add_clause(x, -a)
        return x

    if tag == "imp":
        children = list(elem)
        if len(children) != 2:
            raise ValueError("<imp> must have exactly two children")
        a = tseitin_encode_formula(children[0], cnf)
        b = tseitin_encode_formula(children[1], cnf)
        x = cnf.new_aux("__imp")
        # x <-> (¬a ∨ b)
        # Equivalent to x <-> (a -> b)
        cnf.add_clause(-x, -a, b)
        cnf.add_clause(x, a)
        cnf.add_clause(x, -b)
        return x

    if tag == "eq":
        children = list(elem)
        if len(children) != 2:
            raise ValueError("<eq> must have exactly two children")
        a = tseitin_encode_formula(children[0], cnf)
        b = tseitin_encode_formula(children[1], cnf)
        x = cnf.new_aux("__eq")
        # x <-> (a <-> b)
        # (a<->b) is true when both same.
        cnf.add_clause(-x, -a, b)
        cnf.add_clause(-x, a, -b)
        cnf.add_clause(x, a, b)
        cnf.add_clause(x, -a, -b)
        return x

    raise ValueError(f"Unsupported constraint operator: <{tag}>")


def encode_constraints(xml_root: ET.Element, cnf: CNFBuilder) -> int:
    constraints = xml_root.find("constraints")
    if constraints is None:
        return 0

    count = 0
    for rule in constraints.findall("rule"):
        children = list(rule)
        if len(children) != 1:
            raise ValueError("<rule> must have exactly one top-level formula")
        top_formula = children[0]
        top_var = tseitin_encode_formula(top_formula, cnf)
        cnf.add_clause(top_var)  # rule must hold
        count += 1

    return count


# ---------------------------------------------------------------------
# SAT model counting
# ---------------------------------------------------------------------

def simplify_clauses(
    clauses: List[Tuple[int, ...]],
    assignment: Dict[int, bool]
) -> Optional[List[Tuple[int, ...]]]:
    simplified: List[Tuple[int, ...]] = []

    for clause in clauses:
        new_clause = []
        satisfied = False
        for lit in clause:
            var = abs(lit)
            sign = lit > 0
            if var in assignment:
                if assignment[var] == sign:
                    satisfied = True
                    break
            else:
                new_clause.append(lit)

        if satisfied:
            continue

        if not new_clause:
            return None  # contradiction

        simplified.append(tuple(new_clause))

    return simplified


def unit_propagate(
    clauses: List[Tuple[int, ...]],
    assignment: Dict[int, bool]
) -> Tuple[Optional[List[Tuple[int, ...]]], Dict[int, bool]]:
    changed = True
    clauses_cur = clauses
    assign_cur = dict(assignment)

    while changed:
        changed = False

        # Unit clauses
        units = [c[0] for c in clauses_cur if len(c) == 1]
        if units:
            for lit in units:
                var = abs(lit)
                val = lit > 0
                if var in assign_cur and assign_cur[var] != val:
                    return None, assign_cur
                if var not in assign_cur:
                    assign_cur[var] = val
                    changed = True

            clauses_cur = simplify_clauses(clauses_cur, assign_cur)
            if clauses_cur is None:
                return None, assign_cur
            continue

        # Pure literals
        polarity: Dict[int, Set[bool]] = {}
        for clause in clauses_cur:
            for lit in clause:
                var = abs(lit)
                if var in assign_cur:
                    continue
                polarity.setdefault(var, set()).add(lit > 0)

        pure_assignments = {}
        for var, pols in polarity.items():
            if len(pols) == 1:
                pure_assignments[var] = next(iter(pols))

        if pure_assignments:
            for var, val in pure_assignments.items():
                if var in assign_cur and assign_cur[var] != val:
                    return None, assign_cur
                assign_cur[var] = val
                changed = True

            clauses_cur = simplify_clauses(clauses_cur, assign_cur)
            if clauses_cur is None:
                return None, assign_cur

    return clauses_cur, assign_cur


def choose_branch_var(clauses: List[Tuple[int, ...]], assignment: Dict[int, bool]) -> Optional[int]:
    counter = Counter()
    for clause in clauses:
        for lit in clause:
            var = abs(lit)
            if var not in assignment:
                counter[var] += 1
    if not counter:
        return None
    return counter.most_common(1)[0][0]


def model_count(
    clauses: List[Tuple[int, ...]],
    num_vars: int,
) -> int:
    memo: Dict[Tuple[Tuple[int, bool], Tuple[Tuple[int, ...], ...]], int] = {}

    def rec(
        cur_clauses: List[Tuple[int, ...]],
        assignment: Dict[int, bool]
    ) -> int:
        simplified = simplify_clauses(cur_clauses, assignment)
        if simplified is None:
            return 0

        simplified, assignment = unit_propagate(simplified, assignment)
        if simplified is None:
            return 0

        if not simplified:
            free_vars = num_vars - len(assignment)
            return 2 ** free_vars

        key = (
            tuple(sorted(assignment.items())),
            tuple(sorted(simplified)),
        )
        if key in memo:
            return memo[key]

        var = choose_branch_var(simplified, assignment)
        if var is None:
            free_vars = num_vars - len(assignment)
            return 2 ** free_vars

        total = 0
        for value in (False, True):
            next_assignment = dict(assignment)
            next_assignment[var] = value
            total += rec(simplified, next_assignment)

        memo[key] = total
        return total

    return rec(clauses, {})


# ---------------------------------------------------------------------
# End-to-end analysis
# ---------------------------------------------------------------------

def analyze_feature_model(xml_path: Path) -> Dict[str, int | str]:
    fm_root, xml_root = parse_feature_model(xml_path)
    nodes = collect_nodes(fm_root)

    total_features = len(nodes)
    concrete_features = sum(1 for n in nodes if not n.abstract)

    cnf = CNFBuilder()

    # Register all named nodes before encoding
    for node in nodes:
        cnf.get_var(node.name)

    encode_structure(fm_root, cnf)
    number_of_constraints = encode_constraints(xml_root, cnf)

    # Count only over original named feature variables, not auxiliary Tseitin vars.
    original_feature_var_count = len(cnf.var_to_id)

    # After constraint encoding, aux vars have also been added.
    # We must count over ALL vars in the CNF because auxiliaries are logically fixed by the formula.
    # Since Tseitin variables are constrained to equivalence, each satisfying assignment of original vars
    # corresponds to exactly one satisfying assignment of auxiliary vars.
    total_valid_configurations = model_count(cnf.clauses, cnf.next_id - 1)

    return {
        "SPL Name": xml_path.parent.parent.name,
        "Number of Features in Feature Model": total_features,
        "Number of Concrete Features": concrete_features,
        "Number of Constraints": number_of_constraints,
        "Number of Possible Valid Configurations": total_valid_configurations,
    }


# ---------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------

def write_excel(rows: List[Dict[str, int | str]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Feature Model Analysis"

    headers = [
        "SPL Name",
        "Number of Features in Feature Model",
        "Number of Concrete Features",
        "Number of Constraints",
        "Number of Possible Valid Configurations",
    ]

    ws.append(headers)

    for row in rows:
        ws.append([row[h] for h in headers])

    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Column widths
    widths = {
        "A": 24,
        "B": 34,
        "C": 28,
        "D": 22,
        "E": 38,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Align numeric columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=5):
        for cell in row:
            cell.alignment = center

    ws.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main() -> None:
    rows: List[Dict[str, int | str]] = []

    print("Analyzing feature models...\n")

    for case_name in CASES:
        xml_path = BASE_CASES_DIR / case_name / "configs" / "model.xml"

        if not xml_path.exists():
            print(f"[WARN] Missing: {xml_path}")
            rows.append({
                "SPL Name": case_name,
                "Number of Features in Feature Model": "N/A",
                "Number of Concrete Features": "N/A",
                "Number of Constraints": "N/A",
                "Number of Possible Valid Configurations": "N/A",
            })
            continue

        try:
            print(f"[INFO] Processing {case_name} ...")
            result = analyze_feature_model(xml_path)
            rows.append(result)
            print(
                f"       Features={result['Number of Features in Feature Model']}, "
                f"Concrete={result['Number of Concrete Features']}, "
                f"Constraints={result['Number of Constraints']}, "
                f"ValidConfigs={result['Number of Possible Valid Configurations']}"
            )
        except Exception as exc:
            print(f"[ERROR] {case_name}: {exc}")
            rows.append({
                "SPL Name": case_name,
                "Number of Features in Feature Model": "ERROR",
                "Number of Concrete Features": "ERROR",
                "Number of Constraints": "ERROR",
                "Number of Possible Valid Configurations": "ERROR",
            })

    write_excel(rows, OUTPUT_XLSX)
    print(f"\nDone. Excel written to:\n  {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()